#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel文本替换工具
用于批量替换Excel文件中的文本内容，保持原有格式不变
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import xlrd
import xlwt

from datetime import datetime

# ==================== 配置区域 ====================
# 目标文件夹路径配置
# 设置为空字符串或None时，使用命令行参数指定的路径或当前目录
# 示例: TARGET_FOLDER = r"E:\qyn_game\parseFiles\global\config\test\lang_client"
TARGET_FOLDER = r"E:\qyn_game\parseFiles\global\config\test\lang_client"

# 在这里配置你需要的替换规则
REPLACEMENT_CONFIG = {
    "人才":"能士",
    "知己":"挚友",
    "士人":"势者",
    "农民":"力者",
    "工匠":"韧者",
    "商贾":"智者",
    "武者":"敏者",
    "能力":"修为",
    "头目":"敌首",
    "入驻":"委托",
    "营收":"账收",
    "亲密":"情谊",
    "魅力":"才情",
    "士类":"势类",
    "农类":"力类",
    "工类":"韧类",
    "商类":"智类",
    "武类":"敏类",
    "灵宠":"藏品",
    # 可以添加更多替换规则，格式为 "原文本": "新文本"
}

# 支持的文件扩展名
SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
# ================================================

class ExcelTextReplacer:
    def __init__(self, replacement_config):
        self.replacement_config = replacement_config
        self.total_replacements = 0
        self.processed_files = []
        self.replacement_details = {}
        self.detailed_replacements = []  # 存储详细的替换信息
        self.search_results = []  # 存储搜索结果
        
    def replace_text_in_cell(self, cell_value, file_name, sheet_name, row_idx, col_idx, id_value=""):
        """在单元格文本中进行替换，并记录详细信息"""
        # 统一转换为字符串处理，避免数字类型问题
        if cell_value is None:
            return cell_value, 0

        # 保存原始值
        original_cell_value = cell_value

        # 将所有值转换为字符串进行处理
        cell_value_str = str(cell_value)
        original_value_str = cell_value_str
        replacements_count = 0

        for old_text, new_text in self.replacement_config.items():
            if old_text in cell_value_str:
                count_before = cell_value_str.count(old_text)
                new_cell_value_str = cell_value_str.replace(old_text, new_text)
                replacements_count += count_before

                # 记录替换详情
                if old_text not in self.replacement_details:
                    self.replacement_details[old_text] = 0
                self.replacement_details[old_text] += count_before

                # 记录详细的替换信息
                self.detailed_replacements.append({
                    'file': file_name,
                    'sheet': sheet_name,
                    'row': row_idx + 1,  # 转换为1基索引
                    'col': col_idx + 1,
                    'id': id_value,
                    'before': original_value_str,
                    'after': new_cell_value_str,
                    'old_text': old_text,
                    'new_text': new_text
                })

                cell_value_str = new_cell_value_str

        # 如果有替换，返回字符串；如果没有替换，返回原始值
        if replacements_count > 0:
            return cell_value_str, replacements_count
        else:
            return original_cell_value, replacements_count

    def search_text_in_cell(self, cell_value, search_text, file_name, sheet_name, row_idx, col_idx):
        """在单元格中搜索指定文本"""
        if cell_value is None:
            return False

        # 将所有值转换为字符串进行搜索
        cell_value_str = str(cell_value)

        # 检查是否匹配（支持模糊搜索）
        is_match = self.is_text_match(cell_value_str, search_text)

        if is_match:
            # 获取第1列作为ID（如果存在）
            id_value = ""
            if col_idx == 0:  # 如果当前就是第1列
                id_value = cell_value_str

            self.search_results.append({
                'file': file_name,
                'sheet': sheet_name,
                'row': row_idx + 1,  # 转换为1基索引
                'col': col_idx + 1,
                'id': id_value,
                'content': cell_value_str,
                'search_text': search_text
            })
            return True
        return False

    def is_text_match(self, text, search_text):
        """检查文本是否匹配搜索条件（支持模糊搜索）"""
        if search_text.endswith('*'):
            # 模糊搜索：前缀匹配
            prefix = search_text[:-1]  # 去掉末尾的 *
            return text.startswith(prefix)
        else:
            # 精确搜索：完全匹配（而不是包含匹配）
            return text == search_text

    def copy_cell_style(self, source_cell, target_cell):
        """复制单元格样式"""
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
        
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
        
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )
    
    def process_excel_file(self, file_path):
        """处理单个Excel文件"""
        print(f"正在处理文件: {file_path}")

        file_path_obj = Path(file_path)
        file_extension = file_path_obj.suffix.lower()

        if file_extension == '.xls':
            return self.process_xls_file(file_path)
        elif file_extension == '.xlsx':
            return self.process_xlsx_file(file_path)
        else:
            print(f"不支持的文件格式: {file_extension}")
            return False

    def process_xlsx_file(self, file_path):
        """处理.xlsx文件"""
        try:
            # 读取工作簿
            workbook = openpyxl.load_workbook(file_path)
            file_replacements = 0
            file_name = Path(file_path).name

            print(f"  处理文件: {file_name}")

            # 遍历所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_replacements = 0

                print(f"    工作表: {sheet_name}")

                # 遍历所有行
                for row_idx, row in enumerate(sheet.iter_rows()):
                    id_value = ""
                    # 获取第1列的ID值（如果存在）
                    if len(row) > 0 and row[0].value is not None:
                        id_value = str(row[0].value)

                    # 只处理第1列(ID)和第3列(中文名称)
                    for col_idx in [0, 2]:  # 0基索引，对应第1列和第3列
                        if col_idx < len(row) and row[col_idx].value is not None:
                            cell = row[col_idx]
                            new_value, replacements = self.replace_text_in_cell(
                                cell.value, file_name, sheet_name, row_idx, col_idx, id_value
                            )
                            if replacements > 0:
                                cell.value = new_value
                                sheet_replacements += replacements

                if sheet_replacements > 0:
                    print(f"      完成替换: {sheet_replacements} 处")
                file_replacements += sheet_replacements

            # 直接保存到原文件
            workbook.save(file_path)
            workbook.close()

            self.total_replacements += file_replacements
            self.processed_files.append({
                'file': str(file_path),
                'replacements': file_replacements
            })

            if file_replacements > 0:
                print(f"  ✅ 文件已更新，替换总数: {file_replacements}")
            else:
                print(f"  ⚪ 文件无需更新")

        except Exception as e:
            print(f"处理文件 {file_path} 时出错: {str(e)}")
            return False

        return True

    def process_xls_file(self, file_path):
        """处理.xls文件"""
        try:
            # 读取.xls文件
            workbook_read = xlrd.open_workbook(file_path)
            workbook_write = xlwt.Workbook()
            file_replacements = 0
            file_name = Path(file_path).name

            print(f"  处理文件: {file_name}")

            # 遍历所有工作表
            for sheet_index in range(workbook_read.nsheets):
                sheet_read = workbook_read.sheet_by_index(sheet_index)
                sheet_name = sheet_read.name
                sheet_write = workbook_write.add_sheet(sheet_name)
                sheet_replacements = 0

                print(f"    工作表: {sheet_name}")

                # 遍历所有行
                for row_idx in range(sheet_read.nrows):
                    id_value = ""
                    # 获取第1列的ID值（如果存在）
                    if sheet_read.ncols > 0:
                        id_cell_value = sheet_read.cell_value(row_idx, 0)
                        if id_cell_value:
                            id_value = str(id_cell_value)

                    # 处理所有列，但只对第1列和第3列进行替换检查
                    for col_idx in range(sheet_read.ncols):
                        cell_value = sheet_read.cell_value(row_idx, col_idx)

                        # 只对第1列(ID)和第3列(中文名称)进行替换
                        if col_idx in [0, 2] and cell_value is not None:
                            new_value, replacements = self.replace_text_in_cell(
                                cell_value, file_name, sheet_name, row_idx, col_idx, id_value
                            )
                            if replacements > 0:
                                sheet_write.write(row_idx, col_idx, new_value)
                                sheet_replacements += replacements
                            else:
                                # 保持原始数据类型，避免数字格式问题
                                sheet_write.write(row_idx, col_idx, cell_value)
                        else:
                            # 其他列直接复制，保持原始数据类型
                            sheet_write.write(row_idx, col_idx, cell_value)

                if sheet_replacements > 0:
                    print(f"      完成替换: {sheet_replacements} 处")
                file_replacements += sheet_replacements

            # 直接保存到原文件
            workbook_write.save(str(file_path))

            self.total_replacements += file_replacements
            self.processed_files.append({
                'file': str(file_path),
                'replacements': file_replacements
            })

            if file_replacements > 0:
                print(f"  ✅ 文件已更新，替换总数: {file_replacements}")
            else:
                print(f"  ⚪ 文件无需更新")

        except Exception as e:
            print(f"处理文件 {file_path} 时出错: {str(e)}")
            return False

        return True
    
    def find_excel_files(self, directory):
        """查找目录中的Excel文件"""
        excel_files = []
        directory_path = Path(directory)
        
        if directory_path.is_file():
            if directory_path.suffix.lower() in SUPPORTED_EXTENSIONS:
                excel_files.append(directory_path)
        else:
            for ext in SUPPORTED_EXTENSIONS:
                excel_files.extend(directory_path.glob(f"*{ext}"))
        
        return excel_files

    def search_in_excel_files(self, search_text, directory):
        """在Excel文件中搜索指定文本"""
        excel_files = self.find_excel_files(directory)

        if not excel_files:
            print(f"在路径 '{directory}' 中未找到Excel文件")
            return

        search_type = "模糊搜索" if search_text.endswith('*') else "精确搜索"
        print(f"在 {len(excel_files)} 个Excel文件中{search_type}: '{search_text}'")
        print("="*60)

        for file_path in excel_files:
            self.search_in_single_file(search_text, file_path)

        # 输出搜索结果
        if self.search_results:
            print(f"\n找到 {len(self.search_results)} 个匹配结果:")
            print("="*60)

            # 按文件和行分组，避免重复显示同一行的不同列
            processed_rows = set()

            for result in self.search_results:
                row_key = f"{result['file']}_{result['sheet']}_{result['row']}"

                if row_key not in processed_rows:
                    processed_rows.add(row_key)

                    # 获取ID和中文内容
                    id_content = result['id'] if result['id'] else ""
                    chinese_content = result.get('chinese_content', "")

                    # 输出格式：文件名[工作表名], 行X: ID, 中文内容
                    if chinese_content:
                        print(f"{result['file']}[{result['sheet']}], 行{result['row']}: {id_content}, {chinese_content}")
                    else:
                        print(f"{result['file']}[{result['sheet']}], 行{result['row']}: {id_content}")
        else:
            print(f"\n未找到包含 '{search_text}' 的内容")

    def search_in_single_file(self, search_text, file_path):
        """在单个Excel文件中搜索"""
        file_path_obj = Path(file_path)
        file_extension = file_path_obj.suffix.lower()

        if file_extension == '.xls':
            self.search_in_xls_file(search_text, file_path)
        elif file_extension == '.xlsx':
            self.search_in_xlsx_file(search_text, file_path)

    def search_in_xlsx_file(self, search_text, file_path):
        """在.xlsx文件中搜索"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            file_name = Path(file_path).name

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                for row_idx, row in enumerate(sheet.iter_rows()):
                    # 获取当前行的ID值（第1列）和中文内容（第3列）
                    id_value = ""
                    chinese_value = ""
                    found_match = False

                    if len(row) > 0 and row[0].value is not None:
                        id_value = str(row[0].value)
                    if len(row) > 2 and row[2].value is not None:
                        chinese_value = str(row[2].value)

                    # 在第1列(ID)和第3列(中文名称)中搜索
                    for col_idx in [0, 2]:
                        if col_idx < len(row) and row[col_idx].value is not None:
                            cell_value = str(row[col_idx].value)
                            if self.is_text_match(cell_value, search_text):
                                found_match = True
                                break

                    # 如果找到匹配，添加一个包含完整行信息的结果
                    if found_match:
                        # 确定哪一列包含搜索文本
                        matched_col = 0
                        if id_value and self.is_text_match(id_value, search_text):
                            matched_col = 1
                        elif chinese_value and self.is_text_match(chinese_value, search_text):
                            matched_col = 3

                        # 添加搜索结果，包含完整的行信息
                        self.search_results.append({
                            'file': file_name,
                            'sheet': sheet_name,
                            'row': row_idx + 1,
                            'col': matched_col,
                            'id': id_value,
                            'content': chinese_value if matched_col == 3 else id_value,
                            'chinese_content': chinese_value,  # 总是保存中文内容用于显示
                            'search_text': search_text
                        })

            workbook.close()
        except Exception as e:
            print(f"搜索文件 {file_path} 时出错: {str(e)}")

    def search_in_xls_file(self, search_text, file_path):
        """在.xls文件中搜索"""
        try:
            workbook = xlrd.open_workbook(file_path)
            file_name = Path(file_path).name

            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                sheet_name = sheet.name

                for row_idx in range(sheet.nrows):
                    # 获取当前行的ID值（第1列）和中文内容（第3列）
                    id_value = ""
                    chinese_value = ""
                    found_match = False

                    if sheet.ncols > 0:
                        id_cell_value = sheet.cell_value(row_idx, 0)
                        if id_cell_value:
                            id_value = str(id_cell_value)

                    if sheet.ncols > 2:
                        chinese_cell_value = sheet.cell_value(row_idx, 2)
                        if chinese_cell_value:
                            chinese_value = str(chinese_cell_value)

                    # 在第1列(ID)和第3列(中文名称)中搜索
                    for col_idx in [0, 2]:
                        if col_idx < sheet.ncols:
                            cell_value = sheet.cell_value(row_idx, col_idx)
                            if cell_value:
                                cell_value_str = str(cell_value)
                                if self.is_text_match(cell_value_str, search_text):
                                    found_match = True
                                    break

                    # 如果找到匹配，添加一个包含完整行信息的结果
                    if found_match:
                        # 确定哪一列包含搜索文本
                        matched_col = 0
                        if id_value and self.is_text_match(id_value, search_text):
                            matched_col = 1
                        elif chinese_value and self.is_text_match(chinese_value, search_text):
                            matched_col = 3

                        # 添加搜索结果，包含完整的行信息
                        self.search_results.append({
                            'file': file_name,
                            'sheet': sheet_name,
                            'row': row_idx + 1,
                            'col': matched_col,
                            'id': id_value,
                            'content': chinese_value if matched_col == 3 else id_value,
                            'chinese_content': chinese_value,  # 总是保存中文内容用于显示
                            'search_text': search_text
                        })
        except Exception as e:
            print(f"搜索文件 {file_path} 时出错: {str(e)}")

    def get_chinese_text_by_id(self, search_id, directory=None):
        """根据ID直接获取对应的中文文本（第3列内容）

        Args:
            search_id: 要搜索的ID（如 t_heronew_name500001）
            directory: 搜索目录，如果为None则使用TARGET_FOLDER

        Returns:
            str: 如果找到唯一结果，返回中文文本；如果结果不唯一或未找到，返回None
        """
        if directory is None:
            directory = TARGET_FOLDER

        if not directory:
            return None

        excel_files = self.find_excel_files(directory)
        if not excel_files:
            return None

        matching_results = []

        for file_path in excel_files:
            file_extension = Path(file_path).suffix.lower()

            if file_extension == '.xlsx':
                results = self._search_chinese_in_xlsx(search_id, file_path)
            elif file_extension == '.xls':
                results = self._search_chinese_in_xls(search_id, file_path)
            else:
                continue

            matching_results.extend(results)

        # 检查结果是否唯一
        if len(matching_results) == 1:
            return matching_results[0]['chinese_text']
        else:
            # 结果不唯一或未找到
            return None

    def lookup_field_values(self, excel_file_path, sheet_name, match_column, return_column, search_values):
        """在指定Excel文件的工作表中查找字段值

        Args:
            excel_file_path: Excel文件的绝对路径
            sheet_name: 工作表名称
            match_column: 用于匹配的列名
            return_column: 要返回值的列名
            search_values: 要搜索的值列表

        Returns:
            dict: {search_value: found_value} 的映射字典
        """
        results = {}

        try:
            file_extension = Path(excel_file_path).suffix.lower()

            if file_extension == '.xlsx':
                results = self._lookup_in_xlsx(excel_file_path, sheet_name, match_column, return_column, search_values)
            elif file_extension == '.xls':
                results = self._lookup_in_xls(excel_file_path, sheet_name, match_column, return_column, search_values)

        except Exception as e:
            print(f"查找字段值时出错: {str(e)}")

        return results

    def _lookup_in_xlsx(self, excel_file_path, sheet_name, match_column, return_column, search_values):
        """在.xlsx文件中查找字段值"""
        results = {}

        try:
            workbook = openpyxl.load_workbook(excel_file_path, read_only=True)

            if sheet_name not in workbook.sheetnames:
                print(f"工作表 '{sheet_name}' 不存在于文件 {Path(excel_file_path).name}")
                workbook.close()
                return results

            sheet = workbook[sheet_name]

            # 获取表头行，找到列索引
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

            match_col_idx = None
            return_col_idx = None

            for idx, header in enumerate(header_row):
                if header == match_column and match_col_idx is None:
                    match_col_idx = idx
                elif header == return_column and return_col_idx is None:
                    return_col_idx = idx

            if match_col_idx is None:
                print(f"未找到匹配列 '{match_column}' 在工作表 '{sheet_name}'")
                workbook.close()
                return results

            if return_col_idx is None:
                print(f"未找到返回列 '{return_column}' 在工作表 '{sheet_name}'")
                workbook.close()
                return results

            # 遍历数据行查找匹配值
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > max(match_col_idx, return_col_idx):
                    match_value = row[match_col_idx]
                    return_value = row[return_col_idx]

                    if match_value is not None and return_value is not None:
                        # 统一转换为文本字符串，避免数值类型问题
                        match_str = self._convert_to_text_string(match_value)
                        return_str = self._convert_to_text_string(return_value)
                        if match_str in search_values:
                            results[match_str] = return_str

            workbook.close()

        except Exception as e:
            print(f"读取.xlsx文件时出错: {str(e)}")

        return results

    def _lookup_in_xls(self, excel_file_path, sheet_name, match_column, return_column, search_values):
        """在.xls文件中查找字段值"""
        results = {}

        try:
            workbook = xlrd.open_workbook(excel_file_path)

            sheet_names = workbook.sheet_names()
            if sheet_name not in sheet_names:
                print(f"工作表 '{sheet_name}' 不存在于文件 {Path(excel_file_path).name}")
                return results

            sheet = workbook.sheet_by_name(sheet_name)

            if sheet.nrows == 0:
                return results

            # 获取表头行，找到列索引（只在第1行查找，优先选择列数小的列）
            match_col_idx = None
            return_col_idx = None

            for col_idx in range(sheet.ncols):
                header_value = sheet.cell_value(0, col_idx)
                if header_value == match_column and match_col_idx is None:
                    match_col_idx = col_idx
                elif header_value == return_column and return_col_idx is None:
                    return_col_idx = col_idx

            if match_col_idx is None:
                print(f"未找到匹配列 '{match_column}' 在工作表 '{sheet_name}'")
                return results

            if return_col_idx is None:
                print(f"未找到返回列 '{return_column}' 在工作表 '{sheet_name}'")
                return results

            # 遍历数据行查找匹配值
            for row_idx in range(1, sheet.nrows):
                if match_col_idx < sheet.ncols and return_col_idx < sheet.ncols:
                    match_value = sheet.cell_value(row_idx, match_col_idx)
                    return_value = sheet.cell_value(row_idx, return_col_idx)

                    if match_value and return_value:
                        # 统一转换为文本字符串，避免数值类型问题
                        match_str = self._convert_to_text_string(match_value)
                        return_str = self._convert_to_text_string(return_value)

                        if match_str in search_values:
                            results[match_str] = return_str

        except Exception as e:
            print(f"读取.xls文件时出错: {str(e)}")

        return results

    def _convert_to_text_string(self, value):
        """统一将Excel值转换为文本字符串，避免数值类型问题"""
        if value is None:
            return ""

        if isinstance(value, (int, float)):
            # 对于数值类型，检查是否为整数值
            if isinstance(value, float) and value.is_integer():
                # 浮点数但是整数值，转换为整数字符串
                return str(int(value))
            elif isinstance(value, int):
                # 整数直接转换
                return str(value)
            else:
                # 浮点数保持小数
                return str(value)
        else:
            # 其他类型（字符串等）直接转换并去除首尾空格
            return str(value).strip()

    def _search_chinese_in_xlsx(self, search_id, file_path):
        """在.xlsx文件中搜索ID对应的中文文本"""
        results = []
        try:
            workbook = openpyxl.load_workbook(file_path)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                for row_idx, row in enumerate(sheet.iter_rows()):
                    # 检查第1列是否匹配搜索ID
                    if len(row) > 0 and row[0].value is not None:
                        id_value = str(row[0].value)
                        if id_value == search_id:
                            # 获取第3列的中文内容
                            if len(row) > 2 and row[2].value is not None:
                                chinese_text = str(row[2].value)
                                results.append({
                                    'file': Path(file_path).name,
                                    'sheet': sheet_name,
                                    'row': row_idx + 1,
                                    'chinese_text': chinese_text
                                })

            workbook.close()
        except Exception as e:
            pass

        return results

    def _search_chinese_in_xls(self, search_id, file_path):
        """在.xls文件中搜索ID对应的中文文本"""
        results = []
        try:
            workbook = xlrd.open_workbook(file_path)

            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                sheet_name = sheet.name

                for row_idx in range(sheet.nrows):
                    # 检查第1列是否匹配搜索ID
                    if sheet.ncols > 0:
                        id_cell_value = sheet.cell_value(row_idx, 0)
                        if id_cell_value and str(id_cell_value) == search_id:
                            # 获取第3列的中文内容
                            if sheet.ncols > 2:
                                chinese_cell_value = sheet.cell_value(row_idx, 2)
                                if chinese_cell_value:
                                    chinese_text = str(chinese_cell_value)
                                    results.append({
                                        'file': Path(file_path).name,
                                        'sheet': sheet_name,
                                        'row': row_idx + 1,
                                        'chinese_text': chinese_text
                                    })
        except Exception as e:
            pass

        return results

    def update_language_text_by_id(self, t_id, new_chinese_text, directory=None):
        """更新指定ID的中文文本

        Args:
            t_id: 要更新的ID（如 t_heronew_name500001）
            new_chinese_text: 新的中文文本
            directory: 搜索目录，如果为None则使用TARGET_FOLDER

        Returns:
            bool: 是否更新成功
        """
        if directory is None:
            directory = TARGET_FOLDER

        if not directory:
            return False

        excel_files = self.find_excel_files(directory)
        if not excel_files:
            return False

        # 查找所有匹配的条目
        matching_results = []
        for file_path in excel_files:
            file_extension = Path(file_path).suffix.lower()

            if file_extension == '.xlsx':
                results = self._search_chinese_in_xlsx(t_id, file_path)
            elif file_extension == '.xls':
                results = self._search_chinese_in_xls(t_id, file_path)
            else:
                continue

            matching_results.extend(results)

        if len(matching_results) == 1:
            # 找到唯一匹配项，更新它
            result = matching_results[0]
            print(f"  更新现有条目: {t_id} -> {new_chinese_text}")
            return self._update_language_text_in_file(
                Path(directory) / result['file'],
                result['sheet'],
                result['row'] - 1,  # 转换为0基索引
                new_chinese_text
            )
        elif len(matching_results) == 0:
            # 没找到匹配项，尝试新增
            print(f"  未找到现有条目，尝试新增: {t_id} -> {new_chinese_text}")
            return self._auto_add_new_language_entry(t_id, new_chinese_text, directory)
        else:
            # 找到多个匹配项，无法确定唯一目标
            print(f"  找到 {len(matching_results)} 个匹配项，无法确定唯一更新目标")
            for result in matching_results:
                print(f"    - {result['file']}[{result['sheet']}] 行{result['row']}")
            return False

    def _auto_add_new_language_entry(self, t_id, chinese_text, directory):
        """自动确定目标文件并新增语言条目"""
        # 根据t_id前缀确定目标文件
        target_file, target_sheet = self._determine_language_file_by_id(t_id)

        if not target_file:
            print(f"    无法确定 {t_id} 应该添加到哪个语言文件")
            return False

        target_file_path = Path(directory) / target_file

        if not target_file_path.exists():
            print(f"    目标文件不存在: {target_file}")
            return False

        print(f"    新增到 {target_file}[{target_sheet}]: {t_id} -> {chinese_text}")
        return self.add_new_language_entry(t_id, chinese_text, str(target_file_path), target_sheet)

    def _determine_language_file_by_id(self, t_id):
        """根据t_id确定应该添加到哪个语言文件

        Args:
            t_id: t_*格式的标识符

        Returns:
            tuple: (文件名, 工作表名) 或 (None, None)
        """
        # 根据实际存在的语言文件和t_id前缀确定目标文件
        # 优先使用实际存在的文件

        if t_id.startswith('t_heronew_name') or t_id.startswith('t_hero'):
            # 英雄相关的语言文本，优先使用tableLang.xls
            return ('tableLang.xls', 'functionLang')
        elif t_id.startswith('t_skillnew_name') or t_id.startswith('t_skill'):
            # 技能相关的语言文本，使用tableLang.xls
            return ('tableLang.xls', 'functionLang')
        elif t_id.startswith('t_itemnew_name') or t_id.startswith('t_item'):
            # 物品相关的语言文本，使用tableLang.xls
            return ('tableLang.xls', 'functionLang')
        elif t_id.startswith('t_heroSkillnew_name') or t_id.startswith('t_heroSkill'):
            # 英雄技能相关的语言文本，使用tableLang.xls
            return ('tableLang.xls', 'functionLang')
        elif t_id.startswith('t_act'):
            # 活动相关的语言文本，使用actLang.xls
            return ('actLang.xls', 'actLang')
        elif t_id.startswith('t_client'):
            # 客户端相关的语言文本，使用clientLang.xls
            return ('clientLang.xls', 'clientLang')
        elif t_id.startswith('t_core'):
            # 核心相关的语言文本，使用coreLang.xls
            return ('coreLang.xls', 'coreLang')
        else:
            # 默认添加到tableLang.xls（最通用的语言文件）
            return ('tableLang.xls', 'functionLang')

    def _update_language_text_in_file(self, file_path, sheet_name, row_idx, new_text):
        """在指定文件中更新语言文本"""
        try:
            file_extension = file_path.suffix.lower()

            if file_extension == '.xlsx':
                return self._update_text_in_xlsx(file_path, sheet_name, row_idx, new_text)
            elif file_extension == '.xls':
                return self._update_text_in_xls(file_path, sheet_name, row_idx, new_text)
            else:
                return False

        except Exception as e:
            print(f"更新语言文本时出错: {str(e)}")
            return False

    def _update_text_in_xlsx(self, file_path, sheet_name, row_idx, new_text):
        """在.xlsx文件中更新文本"""
        try:
            workbook = openpyxl.load_workbook(file_path)

            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return False

            sheet = workbook[sheet_name]

            # 更新第3列（索引为2）的文本
            if sheet.max_row > row_idx and sheet.max_column >= 3:
                sheet.cell(row=row_idx + 1, column=3, value=new_text)
                workbook.save(file_path)
                workbook.close()
                print(f"  已更新 {file_path.name}[{sheet_name}] 行{row_idx + 1}")
                return True
            else:
                workbook.close()
                return False

        except Exception as e:
            print(f"更新.xlsx文件时出错: {str(e)}")
            return False

    def _update_text_in_xls(self, file_path, sheet_name, row_idx, new_text):
        """在.xls文件中更新文本（需要重写整个文件）"""
        try:
            # 读取原文件的所有数据
            old_workbook = xlrd.open_workbook(file_path)

            # 创建新的工作簿
            new_workbook = xlwt.Workbook()

            # 复制所有工作表
            for sheet_idx in range(old_workbook.nsheets):
                old_sheet = old_workbook.sheet_by_index(sheet_idx)
                old_sheet_name = old_sheet.name
                new_sheet = new_workbook.add_sheet(old_sheet_name)

                # 复制所有数据
                for r in range(old_sheet.nrows):
                    for c in range(old_sheet.ncols):
                        cell_value = old_sheet.cell_value(r, c)

                        # 如果是目标位置，使用新文本
                        if (old_sheet_name == sheet_name and
                            r == row_idx and c == 2):  # 第3列
                            new_sheet.write(r, c, new_text)
                        else:
                            new_sheet.write(r, c, cell_value)

            # 保存新文件
            new_workbook.save(file_path)
            print(f"  已更新 {file_path.name}[{sheet_name}] 行{row_idx + 1}")
            return True

        except Exception as e:
            print(f"更新.xls文件时出错: {str(e)}")
            return False

    def add_new_language_entry(self, t_id, chinese_text, target_file_path, target_sheet_name):
        """在指定文件中新增语言条目

        Args:
            t_id: 语言ID
            chinese_text: 中文文本
            target_file_path: 目标文件路径
            target_sheet_name: 目标工作表名

        Returns:
            bool: 是否新增成功
        """
        try:
            file_path = Path(target_file_path)
            file_extension = file_path.suffix.lower()

            if file_extension == '.xlsx':
                return self._add_entry_to_xlsx(file_path, target_sheet_name, t_id, chinese_text)
            elif file_extension == '.xls':
                return self._add_entry_to_xls(file_path, target_sheet_name, t_id, chinese_text)
            else:
                return False

        except Exception as e:
            print(f"新增语言条目时出错: {str(e)}")
            return False

    def _add_entry_to_xlsx(self, file_path, sheet_name, t_id, chinese_text):
        """在.xlsx文件中新增条目"""
        try:
            if file_path.exists():
                workbook = openpyxl.load_workbook(file_path)
            else:
                workbook = openpyxl.Workbook()
                # 删除默认工作表
                if 'Sheet' in workbook.sheetnames:
                    del workbook['Sheet']

            # 获取或创建工作表
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
                # 添加表头
                sheet.cell(row=1, column=1, value="ID")
                sheet.cell(row=1, column=2, value="英文")
                sheet.cell(row=1, column=3, value="中文")

            # 找到下一个空行
            next_row = sheet.max_row + 1

            # 添加新条目
            sheet.cell(row=next_row, column=1, value=t_id)
            sheet.cell(row=next_row, column=2, value="")  # 英文列暂时为空
            sheet.cell(row=next_row, column=3, value=chinese_text)

            workbook.save(file_path)
            workbook.close()
            print(f"  已新增条目到 {file_path.name}[{sheet_name}] 行{next_row}")
            return True

        except Exception as e:
            print(f"新增条目到.xlsx文件时出错: {str(e)}")
            return False

    def _add_entry_to_xls(self, file_path, sheet_name, t_id, chinese_text):
        """在.xls文件中新增条目（需要重写整个文件）"""
        try:
            existing_sheets = []
            target_sheet_found = False

            # 如果文件存在，读取现有数据
            if file_path.exists():
                old_workbook = xlrd.open_workbook(file_path)

                for sheet_idx in range(old_workbook.nsheets):
                    old_sheet = old_workbook.sheet_by_index(sheet_idx)
                    old_sheet_name = old_sheet.name

                    # 读取工作表数据
                    sheet_data = []
                    for row_idx in range(old_sheet.nrows):
                        row_data = []
                        for col_idx in range(old_sheet.ncols):
                            cell_value = old_sheet.cell_value(row_idx, col_idx)
                            row_data.append(cell_value)
                        sheet_data.append(row_data)

                    existing_sheets.append((old_sheet_name, sheet_data))

                    if old_sheet_name == sheet_name:
                        target_sheet_found = True

            # 如果目标工作表不存在，创建一个
            if not target_sheet_found:
                # 创建表头
                header_row = [t_id, "", chinese_text]  # ID, 英文, 中文
                existing_sheets.append((sheet_name, [["ID", "英文", "中文"], header_row]))

            # 创建新工作簿
            new_workbook = xlwt.Workbook()

            for sheet_name_iter, sheet_data in existing_sheets:
                new_sheet = new_workbook.add_sheet(sheet_name_iter)

                # 如果是目标工作表，添加新条目
                if sheet_name_iter == sheet_name:
                    # 复制现有数据
                    for row_idx, row_data in enumerate(sheet_data):
                        for col_idx, cell_value in enumerate(row_data):
                            new_sheet.write(row_idx, col_idx, cell_value)

                    # 添加新条目
                    new_row_idx = len(sheet_data)
                    new_sheet.write(new_row_idx, 0, t_id)
                    new_sheet.write(new_row_idx, 1, "")  # 英文列
                    new_sheet.write(new_row_idx, 2, chinese_text)

                    print(f"  已新增条目到 {file_path.name}[{sheet_name}] 行{new_row_idx + 1}")
                else:
                    # 复制其他工作表的数据
                    for row_idx, row_data in enumerate(sheet_data):
                        for col_idx, cell_value in enumerate(row_data):
                            new_sheet.write(row_idx, col_idx, cell_value)

            new_workbook.save(file_path)
            return True

        except Exception as e:
            print(f"新增条目到.xls文件时出错: {str(e)}")
            return False

    def update_cell_value_precisely(self, excel_file_path, sheet_name, row_num, col_name, new_value, arr_pos, arr_type, change_type):
        """精确更新Excel单元格的值

        Args:
            excel_file_path: Excel文件路径
            sheet_name: 工作表名
            row_num: 行号（1基索引，包含表头）
            col_name: 列名
            new_value: 新值
            arr_pos: 数组位置索引
            arr_type: 数组类型 ('single', ',', '[]')
            change_type: 变更类型 ('新增', '删除', '替换')

        Returns:
            bool: 是否更新成功
        """
        try:
            file_path = Path(excel_file_path)
            file_extension = file_path.suffix.lower()

            if file_extension == '.xlsx':
                return self._update_cell_in_xlsx(file_path, sheet_name, row_num, col_name, new_value, arr_pos, arr_type, change_type)
            elif file_extension == '.xls':
                return self._update_cell_in_xls(file_path, sheet_name, row_num, col_name, new_value, arr_pos, arr_type, change_type)
            else:
                print(f"不支持的文件格式: {file_extension}")
                return False

        except Exception as e:
            print(f"精确更新单元格时出错: {str(e)}")
            return False

    def _update_cell_in_xlsx(self, file_path, sheet_name, row_num, col_name, new_value, arr_pos, arr_type, change_type):
        """在.xlsx文件中精确更新单元格"""
        try:
            workbook = openpyxl.load_workbook(file_path)

            if sheet_name not in workbook.sheetnames:
                print(f"工作表 '{sheet_name}' 不存在")
                workbook.close()
                return False

            sheet = workbook[sheet_name]

            # 找到列索引
            col_idx = self._find_column_index(sheet, col_name)
            if col_idx is None:
                print(f"列 '{col_name}' 不存在")
                workbook.close()
                return False

            # 检查行是否存在
            if row_num > sheet.max_row:
                print(f"行 {row_num} 超出范围")
                workbook.close()
                return False

            # 获取当前单元格值
            current_cell = sheet.cell(row=row_num, column=col_idx)
            current_value = current_cell.value or ""

            # 根据数组类型和变更类型更新值
            updated_value = self._apply_array_change(str(current_value), new_value, arr_pos, arr_type, change_type)

            # 更新单元格
            current_cell.value = updated_value

            workbook.save(file_path)
            workbook.close()

            print(f"  已更新 {file_path.name}[{sheet_name}] 行{row_num} 列{col_name}")
            return True

        except Exception as e:
            print(f"更新.xlsx单元格时出错: {str(e)}")
            return False

    def _find_column_index(self, sheet, col_name):
        """查找列名对应的索引"""
        for col_idx, cell in enumerate(sheet[1], 1):  # 第一行是表头
            if cell.value == col_name:
                return col_idx
        return None

    def _apply_array_change(self, current_value, new_value, arr_pos, arr_type, change_type):
        """应用数组变更到当前值

        Args:
            current_value: 当前单元格值
            new_value: 新值
            arr_pos: 数组位置
            arr_type: 数组类型
            change_type: 变更类型

        Returns:
            str: 更新后的值
        """
        if arr_type == 'single':
            # 单个值直接替换
            if change_type == '删除':
                return ""
            else:
                return new_value

        # 解析当前值为数组
        current_items = self._parse_value_to_array(current_value, arr_type)

        # 应用变更
        if change_type == '删除':
            # 删除指定位置的项目
            if 0 <= arr_pos < len(current_items):
                current_items.pop(arr_pos)
        elif change_type == '新增':
            # 在指定位置插入新项目
            if arr_pos <= len(current_items):
                current_items.insert(arr_pos, new_value)
            else:
                current_items.append(new_value)
        elif change_type == '替换':
            # 替换指定位置的项目
            if 0 <= arr_pos < len(current_items):
                current_items[arr_pos] = new_value
            else:
                current_items.append(new_value)

        # 重新组装为原始格式
        return self._format_array_to_string(current_items, arr_type)

    def _parse_value_to_array(self, value, arr_type):
        """将值解析为数组"""
        if not value or not isinstance(value, str):
            return []

        value = value.strip()
        if not value:
            return []

        if arr_type == '[]':
            # [aa, bb] 格式
            if value.startswith('[') and value.endswith(']'):
                inner = value[1:-1].strip()
                if inner:
                    return [item.strip() for item in inner.split(',') if item.strip()]
            return []
        elif arr_type == ',':
            # aa, bb 格式
            return [item.strip() for item in value.split(',') if item.strip()]
        else:
            # single 格式
            return [value] if value else []

    def _format_array_to_string(self, items, arr_type):
        """将数组格式化为字符串"""
        if not items:
            return ""

        if arr_type == '[]':
            return f"[{', '.join(items)}]"
        elif arr_type == ',':
            return ', '.join(items)
        else:  # single
            return items[0] if items else ""

    def _update_cell_in_xls(self, file_path, sheet_name, row_num, col_name, new_value, arr_pos, arr_type, change_type):
        """在.xls文件中精确更新单元格（需要重写整个文件）"""
        try:
            # 读取原文件的所有数据
            old_workbook = xlrd.open_workbook(file_path)

            # 找到目标工作表
            target_sheet = None
            for sheet_idx in range(old_workbook.nsheets):
                sheet = old_workbook.sheet_by_index(sheet_idx)
                if sheet.name == sheet_name:
                    target_sheet = sheet
                    break

            if target_sheet is None:
                print(f"工作表 '{sheet_name}' 不存在")
                return False

            # 找到列索引
            col_idx = None
            for c in range(target_sheet.ncols):
                if target_sheet.cell_value(0, c) == col_name:  # 第0行是表头
                    col_idx = c
                    break

            if col_idx is None:
                print(f"列 '{col_name}' 不存在")
                return False

            # 检查行是否存在
            if row_num - 1 >= target_sheet.nrows:  # row_num是1基索引
                print(f"行 {row_num} 超出范围")
                return False

            # 创建新的工作簿
            new_workbook = xlwt.Workbook()

            # 复制所有工作表
            for sheet_idx in range(old_workbook.nsheets):
                old_sheet = old_workbook.sheet_by_index(sheet_idx)
                old_sheet_name = old_sheet.name
                new_sheet = new_workbook.add_sheet(old_sheet_name)

                # 复制所有数据
                for r in range(old_sheet.nrows):
                    for c in range(old_sheet.ncols):
                        cell_value = old_sheet.cell_value(r, c)

                        # 如果是目标位置，应用变更
                        if (old_sheet_name == sheet_name and
                            r == row_num - 1 and c == col_idx):  # 转换为0基索引

                            # 应用数组变更
                            updated_value = self._apply_array_change(
                                str(cell_value), new_value, arr_pos, arr_type, change_type
                            )
                            new_sheet.write(r, c, updated_value)
                        else:
                            new_sheet.write(r, c, cell_value)

            # 保存新文件
            new_workbook.save(file_path)
            print(f"  已更新 {file_path.name}[{sheet_name}] 行{row_num} 列{col_name}")
            return True

        except Exception as e:
            print(f"更新.xls单元格时出错: {str(e)}")
            return False

    def update_cell_with_multiple_changes(self, excel_file_path, sheet_name, row_num, col_name, cell_changes):
        """处理单元格的多个变更

        Args:
            excel_file_path: Excel文件路径
            sheet_name: 工作表名
            row_num: 行号
            col_name: 列名
            cell_changes: 变更列表，每个变更包含 old_item, new_item, arr_pos, arr_type

        Returns:
            bool: 是否更新成功
        """
        try:
            file_path = Path(excel_file_path)
            file_extension = file_path.suffix.lower()

            if file_extension == '.xlsx':
                return self._update_cell_with_changes_xlsx(file_path, sheet_name, row_num, col_name, cell_changes)
            elif file_extension == '.xls':
                return self._update_cell_with_changes_xls(file_path, sheet_name, row_num, col_name, cell_changes)
            else:
                print(f"不支持的文件格式: {file_extension}")
                return False

        except Exception as e:
            print(f"处理单元格多个变更时出错: {str(e)}")
            return False

    def _update_cell_with_changes_xlsx(self, file_path, sheet_name, row_num, col_name, cell_changes):
        """在.xlsx文件中处理单元格的多个变更"""
        try:
            workbook = openpyxl.load_workbook(file_path)

            if sheet_name not in workbook.sheetnames:
                print(f"工作表 '{sheet_name}' 不存在")
                workbook.close()
                return False

            sheet = workbook[sheet_name]

            # 找到列索引
            col_idx = self._find_column_index(sheet, col_name)
            if col_idx is None:
                print(f"列 '{col_name}' 不存在")
                workbook.close()
                return False

            # 检查行是否存在
            if row_num > sheet.max_row:
                print(f"行 {row_num} 超出范围")
                workbook.close()
                return False

            # 获取当前单元格值
            current_cell = sheet.cell(row=row_num, column=col_idx)
            current_value = current_cell.value or ""

            # 应用所有变更到当前值
            updated_value = self._apply_multiple_changes_to_value(str(current_value), cell_changes)

            # 格式还原处理：只还原加工过程中添加的注释
            final_value = self._restore_processed_annotations(updated_value)

            # 更新单元格
            current_cell.value = final_value

            workbook.save(file_path)
            workbook.close()

            print(f"  已更新 {file_path.name}[{sheet_name}] 行{row_num} 列{col_name}")
            return True

        except Exception as e:
            print(f"更新.xlsx单元格时出错: {str(e)}")
            return False

    def _apply_multiple_changes_to_value(self, current_value, cell_changes):
        """将多个变更应用到单个值

        这个方法需要智能地处理同一个单元格的多个变更，
        比如删除和新增操作需要合并处理
        """
        if not cell_changes:
            return current_value

        # 获取第一个变更的数组类型（假设同一单元格的变更类型一致）
        arr_type = cell_changes[0]['arr_type']

        # 解析当前值为数组
        current_items = self._parse_value_to_array(current_value, arr_type)

        # 按位置分组变更
        changes_by_pos = {}
        for change in cell_changes:
            pos = change['arr_pos']
            if pos not in changes_by_pos:
                changes_by_pos[pos] = []
            changes_by_pos[pos].append(change)

        # 按位置从高到低处理（避免索引变化影响）
        for pos in sorted(changes_by_pos.keys(), reverse=True):
            pos_changes = changes_by_pos[pos]

            # 处理该位置的变更
            for change in pos_changes:
                old_item = change['old_item']
                new_item = change['new_item']

                if old_item and new_item:
                    # 替换操作
                    if 0 <= pos < len(current_items):
                        current_items[pos] = new_item
                elif new_item and not old_item:
                    # 新增操作
                    if pos <= len(current_items):
                        current_items.insert(pos, new_item)
                    else:
                        current_items.append(new_item)
                elif old_item and not new_item:
                    # 删除操作
                    if 0 <= pos < len(current_items):
                        current_items.pop(pos)

        # 重新组装为原始格式
        return self._format_array_to_string(current_items, arr_type)

    def _restore_processed_annotations(self, value):
        """格式还原处理：只还原加工过程中添加的注释

        这个方法只处理两种加工过程中添加的注释格式：
        1. t_*{中文} - 语言文本注释
        2. 数字{中文} - ID关联注释

        不处理原本就存在的配置数据，如：力量{+10}

        Args:
            value: 包含注释的值

        Returns:
            str: 还原后的值
        """
        if not value or not isinstance(value, str):
            return value

        import re

        # 只处理 t_*{中文} 格式（语言文本注释）
        t_pattern = r't_([a-zA-Z0-9_]+)\{[^}]*\}'
        processed_value = re.sub(t_pattern, r't_\1', value)

        # 只处理 数字{中文} 格式（ID关联注释）
        # 注意：这里要确保是纯数字开头的，避免误处理如 "力量100{+10}" 这样的配置
        id_pattern = r'\b(\d+)\{[^}]*\}'
        processed_value = re.sub(id_pattern, r'\1', processed_value)

        return processed_value

    def _update_cell_with_changes_xls(self, file_path, sheet_name, row_num, col_name, cell_changes):
        """在.xls文件中处理单元格的多个变更（需要重写整个文件）"""
        try:
            # 读取原文件的所有数据
            old_workbook = xlrd.open_workbook(file_path)

            # 找到目标工作表
            target_sheet = None
            for sheet_idx in range(old_workbook.nsheets):
                sheet = old_workbook.sheet_by_index(sheet_idx)
                if sheet.name == sheet_name:
                    target_sheet = sheet
                    break

            if target_sheet is None:
                print(f"工作表 '{sheet_name}' 不存在")
                return False

            # 找到列索引
            col_idx = None
            for c in range(target_sheet.ncols):
                if target_sheet.cell_value(0, c) == col_name:  # 第0行是表头
                    col_idx = c
                    break

            if col_idx is None:
                print(f"列 '{col_name}' 不存在")
                return False

            # 检查行是否存在
            if row_num - 1 >= target_sheet.nrows:  # row_num是1基索引
                print(f"行 {row_num} 超出范围")
                return False

            # 获取当前单元格值
            current_value = target_sheet.cell_value(row_num - 1, col_idx)  # 转换为0基索引

            # 应用所有变更
            updated_value = self._apply_multiple_changes_to_value(str(current_value), cell_changes)

            # 格式还原处理：只还原加工过程中添加的注释
            final_value = self._restore_processed_annotations(updated_value)

            # 创建新的工作簿
            new_workbook = xlwt.Workbook()

            # 复制所有工作表
            for sheet_idx in range(old_workbook.nsheets):
                old_sheet = old_workbook.sheet_by_index(sheet_idx)
                old_sheet_name = old_sheet.name
                new_sheet = new_workbook.add_sheet(old_sheet_name)

                # 复制所有数据
                for r in range(old_sheet.nrows):
                    for c in range(old_sheet.ncols):
                        cell_value = old_sheet.cell_value(r, c)

                        # 如果是目标位置，使用格式还原后的值
                        if (old_sheet_name == sheet_name and
                            r == row_num - 1 and c == col_idx):  # 转换为0基索引
                            new_sheet.write(r, c, final_value)
                        else:
                            new_sheet.write(r, c, cell_value)

            # 保存新文件
            new_workbook.save(file_path)
            print(f"  已更新 {file_path.name}[{sheet_name}] 行{row_num} 列{col_name}")
            return True

        except Exception as e:
            print(f"更新.xls单元格时出错: {str(e)}")
            return False

    def get_id_for_row(self, file_name, sheet_name, row_idx):
        """获取指定行的ID值（第1列）"""
        # 从搜索结果中查找对应行的ID
        for result in self.search_results:
            if (result['file'] == file_name and
                result['sheet'] == sheet_name and
                result['row'] == row_idx + 1 and
                result['col'] == 1):
                return result['content']

        # 如果没找到，尝试重新读取文件获取ID
        try:
            file_path = Path(file_name)
            if file_path.suffix.lower() == '.xlsx':
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook[sheet_name]
                rows = list(sheet.iter_rows())
                if row_idx < len(rows) and len(rows[row_idx]) > 0:
                    id_value = rows[row_idx][0].value
                    workbook.close()
                    return str(id_value) if id_value is not None else ""
            elif file_path.suffix.lower() == '.xls':
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_name(sheet_name)
                if row_idx < sheet.nrows and sheet.ncols > 0:
                    id_value = sheet.cell_value(row_idx, 0)
                    return str(id_value) if id_value else ""
        except:
            pass

        return ""

    def print_summary(self):
        """打印处理总结"""
        print("\n" + "="*80)
        print("处理总结")
        print("="*80)
        print(f"处理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"处理文件数量: {len(self.processed_files)}")
        print(f"总替换次数: {self.total_replacements}")

        print("\n替换规则:")
        for old_text, new_text in self.replacement_config.items():
            count = self.replacement_details.get(old_text, 0)
            print(f"  '{old_text}' → '{new_text}' (替换了 {count} 次)")

        print("\n处理的文件:")
        for file_info in self.processed_files:
            print(f"  文件: {file_info['file']}")
            print(f"  替换次数: {file_info['replacements']}")

        # 输出详细的替换信息
        if self.detailed_replacements:
            print("\n" + "="*80)
            print("详细替换记录")
            print("="*80)

            # 按文件分组显示
            current_file = ""
            for replacement in self.detailed_replacements:
                if replacement['file'] != current_file:
                    current_file = replacement['file']
                    print(f"\n📁 {current_file}:")

                # 格式化输出替换信息
                col_name = "ID" if replacement['col'] == 1 else "中文名称"
                print(f"  [{replacement['sheet']}], 行{replacement['row']}({col_name}): {replacement['id']},{replacement['col']},{replacement['before']} -> {replacement['id']},{replacement['col']},{replacement['after']}")

        print("\n" + "="*80)

def main():
    # 创建替换器实例
    replacer = ExcelTextReplacer(REPLACEMENT_CONFIG)

    # 确定工作路径：优先使用配置的目标文件夹
    if TARGET_FOLDER and TARGET_FOLDER.strip():
        work_path = TARGET_FOLDER
        print(f"使用配置的目标文件夹: {work_path}")
    else:
        # 简单的命令行参数处理
        if len(sys.argv) >= 2:
            first_arg = sys.argv[1]

            # 检查第一个参数是否是搜索文本（包含引号或不包含.xls/.xlsx扩展名）
            if (first_arg.startswith('"') and first_arg.endswith('"')) or \
               (not first_arg.endswith('.xls') and not first_arg.endswith('.xlsx') and not first_arg == '.'):
                # 搜索模式
                search_text = first_arg.strip('"')
                work_path = sys.argv[2] if len(sys.argv) >= 3 else '.'
                print("Excel文本搜索工具")
                print("="*40)
                print("支持模糊搜索：在搜索文本末尾添加 * 号进行前缀匹配")
                print("例如：'t_hero_getway*' 可搜索所有以 t_hero_getway 开头的文本")
                print("="*40)
                print(f"搜索路径: {work_path}")
                replacer.search_in_excel_files(search_text, work_path)
                return
            else:
                # 替换模式，第一个参数是路径
                work_path = first_arg
        else:
            # 默认当前目录
            work_path = '.'

    # 处理搜索模式（当配置了目标文件夹时）
    if len(sys.argv) >= 2 and TARGET_FOLDER and TARGET_FOLDER.strip():
        first_arg = sys.argv[1]
        # 检查第一个参数是否是搜索文本
        if (first_arg.startswith('"') and first_arg.endswith('"')) or \
           (not first_arg.endswith('.xls') and not first_arg.endswith('.xlsx') and not first_arg == '.'):
            # 搜索模式
            search_text = first_arg.strip('"')
            print("Excel文本搜索工具")
            print("="*40)
            print("支持模糊搜索：在搜索文本末尾添加 * 号进行前缀匹配")
            print("例如：'t_hero_getway*' 可搜索所有以 t_hero_getway 开头的文本")
            print("="*40)
            print(f"搜索路径: {work_path}")
            replacer.search_in_excel_files(search_text, work_path)
            return

    print("Excel文本替换工具")
    print("="*40)
    print("当前替换配置:")
    for old_text, new_text in REPLACEMENT_CONFIG.items():
        print(f"  '{old_text}' → '{new_text}'")
    print(f"工作路径: {work_path}")
    print()

    # 查找Excel文件
    excel_files = replacer.find_excel_files(work_path)

    if not excel_files:
        print(f"在路径 '{work_path}' 中未找到Excel文件")
        return

    print(f"找到 {len(excel_files)} 个Excel文件:")
    for file_path in excel_files:
        print(f"  {file_path}")
    print()

    # 处理文件
    for file_path in excel_files:
        replacer.process_excel_file(file_path)
        print()

    # 打印总结
    replacer.print_summary()

if __name__ == "__main__":
    main()
