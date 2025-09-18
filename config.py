#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel工作表导出为CSV工具
用于将指定Excel文件的指定工作表导出为CSV格式
使用方法: py config.py hero[hero]
"""

import sys
from pathlib import Path
import pandas as pd
import openpyxl
import xlrd
import xlwt
import re
import multiprocessing
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

import shutil
from wcwidth import wcswidth
import difflib

# ==================== 配置区域 ====================
# 目标文件夹路径
TARGET_FOLDER = r"E:\qyn_game\parseFiles\global\config\test"

# 输出文件夹名称（在目标文件夹下创建）
OUTPUT_FOLDER = "xls"

# 基线备份文件夹名称（在当前工作目录下创建）
BASE_FOLDER = "xls_base"

# 支持的文件扩展名
SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']

# ==================== 预预处理配置区域 ====================
# 自定义字段关联配置
# 格式: "源表[源工作表], 源列名": "目标表[目标工作表], 匹配列名, 返回列名"
# 示例: "hero[hero], 技能-初始资质": "heroSkill[heroskill], 技能id, 名称"
PRE_PROCESSING_CONFIG = {
    "hero[hero], 技能-初始资质": "heroSkill[heroskill], 技能id, 名称",
    "hero[hero], 技能-商铺": "heroSkill[heroskill], 技能id, 名称",
    "hero[hero], 潜能技能": "heroSkill[heroskill], 技能id, 名称",
    "hero[hero], 光环": "heroSkill[heroskill], 技能id, 名称",
    "wife[wif], 门客缘分": "hero[hero], 人才ID, 名字",
    "wife[wif], 技能": "wife[老婆技能], 序号, 技能名",
    # 可以添加更多关联配置
    # "表名[工作表], 列名": "目标表[目标工作表], 匹配列, 返回列",
}
# ================================================

class ExcelToCSVConverter:
    def __init__(self, target_folder, output_folder):
        self.target_folder = Path(target_folder)
        # 输出文件夹在当前工作目录下，而不是目标文件夹下
        self.output_folder = Path.cwd() / output_folder

        # 确保输出文件夹存在
        self.output_folder.mkdir(exist_ok=True)

        # 基线备份文件夹
        self.base_folder = Path.cwd() / BASE_FOLDER
        self.base_folder.mkdir(exist_ok=True)


    def parse_command(self, command):
        """解析命令行参数，提取文件名和工作表名"""
        if '[' not in command or ']' not in command:
            raise ValueError("命令格式错误，应为: filename[sheetname]")

        # 分离文件名和工作表名
        file_part, sheet_part = command.split('[', 1)
        sheet_name = sheet_part.rstrip(']')

        if not file_part or not sheet_name:
            raise ValueError("文件名或工作表名不能为空")

        # 自动添加.xls扩展名
        filename = f"{file_part.strip()}.xls"

        return filename, sheet_name.strip()

    def find_excel_file(self, filename):
        """在目标文件夹中查找Excel文件"""
        file_path = self.target_folder / filename

        if file_path.exists():
            return file_path

        # 如果没有找到，尝试不同的扩展名
        name_without_ext = file_path.stem
        for ext in SUPPORTED_EXTENSIONS:
            test_path = self.target_folder / f"{name_without_ext}{ext}"
            if test_path.exists():
                return test_path

        return None

    def read_excel_sheet(self, file_path, sheet_name):
        """读取Excel文件的指定工作表"""
        file_extension = file_path.suffix.lower()

        try:
            if file_extension == '.xlsx':
                return self.read_xlsx_sheet(file_path, sheet_name)
            elif file_extension == '.xls':
                return self.read_xls_sheet(file_path, sheet_name)
            else:
                raise ValueError(f"不支持的文件格式: {file_extension}")
        except Exception as e:
            raise Exception(f"读取Excel文件失败: {str(e)}")

    def read_xlsx_sheet(self, file_path, sheet_name):
        """读取.xlsx文件的指定工作表"""
        try:
            # 首先检查工作表是否存在
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name not in workbook.sheetnames:
                available_sheets = ', '.join(workbook.sheetnames)
                raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {available_sheets}")
            workbook.close()

            # 使用pandas读取指定工作表
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            return df
        except Exception as e:
            raise Exception(f"读取.xlsx文件失败: {str(e)}")

    def read_xls_sheet(self, file_path, sheet_name):
        """读取.xls文件的指定工作表"""
        try:
            # 首先检查工作表是否存在
            workbook = xlrd.open_workbook(file_path)
            sheet_names = workbook.sheet_names()
            if sheet_name not in sheet_names:
                available_sheets = ', '.join(sheet_names)
                raise ValueError(f"工作表 '{sheet_name}' 不存在。可用工作表: {available_sheets}")

            # 使用pandas读取指定工作表
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
            return df
        except Exception as e:
            raise Exception(f"读取.xls文件失败: {str(e)}")

    def find_t_strings(self, text):
        """查找文本中所有的t_*字符串，包括{}内的t_*字符串"""
        if pd.isna(text) or not isinstance(text, str):
            return []

        t_strings = []

        # 方法1: 查找普通的t_*字符串（按逗号分割）
        parts = [part.strip() for part in text.split(',')]
        for part in parts:
            # 检查每个部分是否是完整的t_*字符串
            if re.match(r'^t_[a-zA-Z0-9_]+$', part):
                t_strings.append(part)

        # 方法2: 查找{}内的t_*字符串，如 1001{t_heroSkillnew_name1001}
        brace_pattern = r'\{(t_[a-zA-Z0-9_]+)\}'
        t_strings_in_braces = re.findall(brace_pattern, text)
        t_strings.extend(t_strings_in_braces)

        # 方法3: 查找所有独立的t_*字符串（不在{}内的）
        # 这个用于处理可能遗漏的情况
        all_t_pattern = r't_[a-zA-Z0-9_]+'
        all_t_strings = re.findall(all_t_pattern, text)

        # 过滤掉已经在{}内的t_*字符串，避免重复处理
        for t_str in all_t_strings:
            # 检查这个t_*字符串是否在{}内
            if f'{{{t_str}}}' not in text:
                t_strings.append(t_str)

        return list(set(t_strings))  # 去重

    def search_chinese_text(self, t_string):
        """直接调用go.py的方法获取t_string对应的中文文本"""
        try:
            # 导入go.py模块
            import sys
            sys.path.append(str(Path.cwd()))
            from go import ExcelTextReplacer, TARGET_FOLDER

            # 创建ExcelTextReplacer实例
            replacer = ExcelTextReplacer({})  # 空的替换配置，因为我们只是用来搜索

            # 直接调用新的方法获取中文文本
            chinese_text = replacer.get_chinese_text_by_id(t_string, TARGET_FOLDER)

            return chinese_text

        except Exception:
            # 如果出现任何错误，返回None
            return None

    def search_chinese_text_batch(self, t_strings, max_workers=None):
        """并发批量搜索t_string对应的中文文本"""
        if max_workers is None:
            # 根据CPU核心数和任务数量动态调整线程数
            cpu_count = multiprocessing.cpu_count()
            # 使用CPU核心数的2-3倍，但不超过任务数量，最少4个线程
            max_workers = min(max(cpu_count * 3, 4), len(t_strings), 32)

        print(f"使用 {max_workers} 个线程并发搜索...")

        results = {}
        completed_count = 0
        total_count = len(t_strings)
        lock = threading.Lock()

        def search_single(t_string):
            nonlocal completed_count
            chinese_text = self.search_chinese_text(t_string)

            with lock:
                completed_count += 1
                if chinese_text:
                    print(f"  [{completed_count}/{total_count}] {t_string} -> {chinese_text}")
                else:
                    print(f"  [{completed_count}/{total_count}] {t_string} -> 未找到")

            return t_string, chinese_text

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任务
            future_to_string = {executor.submit(search_single, t_string): t_string
                              for t_string in t_strings}

            # 收集结果
            for future in as_completed(future_to_string):
                try:
                    t_string, chinese_text = future.result()
                    results[t_string] = chinese_text
                except Exception as e:
                    t_string = future_to_string[future]
                    print(f"搜索 {t_string} 时出错: {str(e)}")
                    results[t_string] = None

        return results

    def parse_ids_from_value(self, value):
        """从值中解析出ID列表，支持多种格式"""
        if pd.isna(value) or not isinstance(value, str):
            return []

        # 移除空白字符
        value = value.strip()
        if not value:
            return []

        ids = []

        # 处理数组格式 [1001, 1002] 或 [1001,1002]
        if value.startswith('[') and value.endswith(']'):
            # 移除方括号
            inner_value = value[1:-1].strip()
            if inner_value:
                # 按逗号分割
                parts = [part.strip() for part in inner_value.split(',')]
                ids.extend([part for part in parts if part])
        else:
            # 处理逗号分隔格式 1001, 1002 或单个值 1001
            parts = [part.strip() for part in value.split(',')]
            ids.extend([part for part in parts if part])

        return ids

    def pre_preprocess_dataframe(self, df, current_table_sheet):
        """预预处理DataFrame，根据配置进行字段关联查找"""
        if not PRE_PROCESSING_CONFIG:
            print("未配置预预处理规则，跳过预预处理")
            return df

        print("正在进行预预处理，处理自定义字段关联...")

        # 导入go.py模块
        try:
            import sys
            sys.path.append(str(Path.cwd()))
            from go import ExcelTextReplacer
        except Exception as e:
            print(f"导入go.py模块失败: {str(e)}")
            return df

        # 创建查找器实例
        replacer = ExcelTextReplacer({})
        df_processed = df.copy()

        # 处理每个配置规则
        for source_config, target_config in PRE_PROCESSING_CONFIG.items():
            try:
                # 解析源配置: "hero[hero], 技能-初始资质"
                source_parts = [part.strip() for part in source_config.split(',')]
                if len(source_parts) != 2:
                    print(f"源配置格式错误: {source_config}")
                    continue

                source_table_sheet = source_parts[0]  # "hero[hero]"
                source_column = source_parts[1]       # "技能-初始资质"

                # 检查是否匹配当前处理的表和工作表
                if source_table_sheet != current_table_sheet:
                    continue

                # 解析目标配置: "heroSkill[heroskill], 技能id, 名称"
                target_parts = [part.strip() for part in target_config.split(',')]
                if len(target_parts) != 3:
                    print(f"目标配置格式错误: {target_config}")
                    continue

                target_table_sheet = target_parts[0]  # "heroSkill[heroskill]"
                match_column = target_parts[1]        # "技能id"
                return_column = target_parts[2]       # "名称"

                # 解析目标表信息
                if '[' not in target_table_sheet or ']' not in target_table_sheet:
                    print(f"目标表格式错误: {target_table_sheet}")
                    continue

                target_file_part, target_sheet_part = target_table_sheet.split('[', 1)
                target_sheet_name = target_sheet_part.rstrip(']')
                target_filename = f"{target_file_part.strip()}.xls"

                # 构建目标文件的绝对路径
                target_file_path = Path(TARGET_FOLDER) / target_filename

                if not target_file_path.exists():
                    # 尝试其他扩展名
                    for ext in SUPPORTED_EXTENSIONS:
                        test_path = Path(TARGET_FOLDER) / f"{target_file_part.strip()}{ext}"
                        if test_path.exists():
                            target_file_path = test_path
                            break
                    else:
                        print(f"未找到目标文件: {target_filename}")
                        continue

                print(f"处理字段关联: {source_column} -> {target_file_path.name}[{target_sheet_name}]")

                # 检查源列是否存在
                if source_column not in df_processed.columns:
                    print(f"源列 '{source_column}' 不存在")
                    continue

                # 收集所有需要查找的ID
                all_ids = set()
                for idx, cell_value in df_processed[source_column].items():
                    ids = self.parse_ids_from_value(cell_value)
                    all_ids.update(ids)

                if not all_ids:
                    print(f"在列 '{source_column}' 中未找到任何ID")
                    continue

                print(f"找到 {len(all_ids)} 个唯一ID，正在查找对应值...")

                # 使用并发优化的查找方法
                lookup_results = self._lookup_field_values_concurrent(
                    replacer,
                    str(target_file_path),
                    target_sheet_name,
                    match_column,
                    return_column,
                    list(all_ids)
                )

                found_count = len(lookup_results)
                print(f"成功找到 {found_count}/{len(all_ids)} 个ID的对应值")

                # 替换DataFrame中的内容
                for idx, cell_value in df_processed[source_column].items():
                    if pd.notna(cell_value) and isinstance(cell_value, str):
                        ids = self.parse_ids_from_value(cell_value)
                        if ids:
                            # 构建新值，格式: id{对应值}
                            new_parts = []
                            for id_val in ids:
                                if id_val in lookup_results:
                                    new_parts.append(f"{id_val}{{{lookup_results[id_val]}}}")
                                else:
                                    new_parts.append(id_val)  # 保持原值

                            # 根据原格式重新组装
                            original_value = str(cell_value).strip()
                            if original_value.startswith('[') and original_value.endswith(']'):
                                # 保持数组格式
                                df_processed.loc[idx, source_column] = f"[{', '.join(new_parts)}]"
                            else:
                                # 保持逗号分隔格式
                                df_processed.loc[idx, source_column] = ', '.join(new_parts)

                print(f"完成字段 '{source_column}' 的关联处理")

            except Exception as e:
                print(f"处理配置 '{source_config}' 时出错: {str(e)}")
                continue

        print("预预处理完成")
        return df_processed

    def _lookup_field_values_concurrent(self, replacer, file_path, sheet_name, match_column, return_column, search_values):
        """并发优化的字段值查找方法"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import multiprocessing

        if not search_values:
            return {}

        # 动态调整线程数
        cpu_count = multiprocessing.cpu_count()
        max_workers = min(max(cpu_count * 2, 4), len(search_values), 16)  # 预预处理用较少线程

        print(f"使用 {max_workers} 个线程并发查找字段值...")

        # 将搜索值分批处理
        batch_size = max(1, len(search_values) // max_workers)
        batches = [search_values[i:i + batch_size] for i in range(0, len(search_values), batch_size)]

        def lookup_batch(batch_values):
            """查找一批值"""
            return replacer.lookup_field_values(file_path, sheet_name, match_column, return_column, batch_values)

        # 并发执行查找
        all_results = {}
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有批次任务
            future_to_batch = {executor.submit(lookup_batch, batch): batch for batch in batches}

            # 收集结果
            completed_count = 0
            for future in as_completed(future_to_batch):
                batch = future_to_batch[future]
                try:
                    batch_results = future.result()
                    all_results.update(batch_results)
                    completed_count += len(batch)
                    print(f"  已完成 {completed_count}/{len(search_values)} 个ID的查找")
                except Exception as e:
                    print(f"  批次查找失败: {str(e)}")

        return all_results

    def preprocess_dataframe(self, df):
        """预处理DataFrame，将t_*字符串替换为t_*{中文}格式"""
        print("正在进行预处理，识别并查找t_*字符串...")

        # 收集所有需要查找的t_*字符串
        all_t_strings = set()

        for col in df.columns:
            for idx, cell_value in df[col].items():
                if pd.notna(cell_value) and isinstance(cell_value, str):
                    t_strings = self.find_t_strings(cell_value)
                    all_t_strings.update(t_strings)

        if not all_t_strings:
            print("未找到任何t_*字符串，跳过预处理")
            return df

        print(f"找到 {len(all_t_strings)} 个唯一的t_*字符串，正在并发查找对应中文...")

        # 并发批量查找中文文本
        chinese_results = self.search_chinese_text_batch(list(all_t_strings))

        # 构建替换映射
        t_string_map = {}
        found_count = 0
        for t_string, search_result in chinese_results.items():
            if search_result:
                # search_result 现在只包含中文内容
                t_string_map[t_string] = f"{t_string}{{{search_result}}}"
                found_count += 1
            else:
                t_string_map[t_string] = t_string  # 保持原样（搜索结果不唯一或未找到）

        # 替换DataFrame中的内容
        print("正在替换DataFrame中的内容...")
        df_processed = df.copy()

        for col in df_processed.columns:
            for idx, cell_value in df_processed[col].items():
                if pd.notna(cell_value) and isinstance(cell_value, str):
                    new_value = cell_value

                    # 方法1: 直接替换完整的t_*字符串（按逗号分割）
                    parts = [part.strip() for part in new_value.split(',')]
                    new_parts = []

                    for part in parts:
                        # 检查这个部分是否是完整的t_*字符串
                        if part in t_string_map:
                            new_parts.append(t_string_map[part])
                        else:
                            new_parts.append(part)

                    new_value = ', '.join(new_parts)

                    # 方法2: 替换{}内的t_*字符串
                    import re
                    def replace_t_in_braces(match):
                        t_string = match.group(1)  # 获取{}内的t_*字符串
                        if t_string in t_string_map:
                            return '{' + t_string_map[t_string] + '}'
                        else:
                            return match.group(0)  # 保持原样

                    # 使用正则表达式替换{}内的t_*字符串
                    new_value = re.sub(r'\{(t_[a-zA-Z0-9_]+)\}', replace_t_in_braces, new_value)

                    df_processed.loc[idx, col] = new_value

        print(f"预处理完成，共找到 {found_count}/{len(all_t_strings)} 个t_*字符串的中文对应")
        print(f"替换了 {found_count} 个t_*字符串为带中文的格式")
        return df_processed

    def save_to_csv(self, dataframe, output_filename):
        """将DataFrame保存为CSV文件"""
        output_path = self.output_folder / output_filename

        try:
            # 保存为CSV，使用UTF-8编码，禁用引号转义
            dataframe.to_csv(output_path, index=False, encoding='utf-8-sig',
                           quoting=1, escapechar=None)  # quoting=1 表示 QUOTE_ALL
            return output_path
        except Exception as e:
            raise Exception(f"保存CSV文件失败: {str(e)}")


    def get_display_width(self, text):
        """获取文本的显示宽度（考虑中文字符）"""
        if not text:
            return 0
        width = wcswidth(str(text))
        return width if width is not None else len(str(text))

    def truncate_text(self, text, max_width):
        """截断文本到指定显示宽度"""
        if not text:
            return ""
        text = str(text)
        if self.get_display_width(text) <= max_width:
            return text

        # 逐字符截断直到符合宽度
        result = ""
        for char in text:
            if self.get_display_width(result + char + "...") > max_width:
                return result + "..."
            result += char
        return result

    def parse_array_value(self, value):
        """解析数组值，统一将所有值视为数组处理

        支持格式：
        - 单个值: 'aa' -> ['aa']
        - 逗号分隔: 'aa, bb' -> ['aa', 'bb']
        - 数组格式: '[aa, bb]' -> ['aa', 'bb']

        Returns:
            tuple: (items_list, array_type)
            - items_list: 解析后的项目列表
            - array_type: 原始格式类型 ('single', ',', '[]')
        """
        if not value or not isinstance(value, str):
            return [], 'single'

        value = value.strip()
        if not value:
            return [], 'single'

        # 处理 [aa, bb] 格式
        if value.startswith('[') and value.endswith(']'):
            inner = value[1:-1].strip()
            if inner:
                items = [item.strip() for item in inner.split(',')]
                return [item for item in items if item], '[]'
            else:
                return [], '[]'

        # 处理逗号分隔格式 aa, bb
        elif ',' in value:
            items = [item.strip() for item in value.split(',')]
            return [item for item in items if item], ','

        # 处理单个值 aa
        else:
            return [value], 'single'

    def compare_values_with_diff(self, old_val, new_val):
        """使用difflib智能比较两个值，返回变更描述

        统一将所有值视为数组处理，返回6元素元组格式：
        (change_type, arr_pos, old_item, new_item, original_arr_pos, arr_type)
        """
        # 解析为数组格式（统一处理）
        old_items, old_type = self.parse_array_value(old_val)
        new_items, new_type = self.parse_array_value(new_val)

        # 使用新的数组类型（优先使用new_type，如果为空则使用old_type）
        arr_type = new_type if new_items else old_type

        return self._compare_array_items_unified(old_items, new_items, arr_type)

    def _compare_array_items_unified(self, old_items, new_items, arr_type):
        """统一的数组项目比较方法

        Args:
            old_items: 旧值的项目列表
            new_items: 新值的项目列表
            arr_type: 数组类型 ('single', ',', '[]')

        Returns:
            list: 变更记录列表，格式为 (change_type, arr_pos, old_item, new_item, original_arr_pos, arr_type)
        """
        if old_items == new_items:
            return []

        differ = difflib.SequenceMatcher(None, old_items, new_items)
        opcodes = differ.get_opcodes()

        changes = []
        for tag, i1, i2, j1, j2 in opcodes:
            if tag == 'delete':
                deleted_items = old_items[i1:i2]
                for idx, item in enumerate(deleted_items):
                    changes.append(('删除', i1 + idx, item, None, i1 + idx, arr_type))

            elif tag == 'insert':
                inserted_items = new_items[j1:j2]
                for idx, item in enumerate(inserted_items):
                    changes.append(('新增', j1 + idx, None, item, j1 + idx, arr_type))

            elif tag == 'replace':
                old_part = old_items[i1:i2]
                new_part = new_items[j1:j2]
                # 对于替换，我们将其分解为删除+新增
                for idx, item in enumerate(old_part):
                    changes.append(('删除', i1 + idx, item, None, i1 + idx, arr_type))
                for idx, item in enumerate(new_part):
                    changes.append(('新增', j1 + idx, None, item, j1 + idx, arr_type))

        return changes

    def show_diff_with_baseline(self, csv_file_path):
        """与基线备份进行比对并打印变更摘要"""
        try:
            csv_path = Path(csv_file_path)
            base_path = self.base_folder / csv_path.name
            if not base_path.exists():
                print(f"  无基线备份，跳过比对")
                return

            # 读取当前与基线
            df_curr = pd.read_csv(csv_path, encoding='utf-8-sig')
            df_base = pd.read_csv(base_path, encoding='utf-8-sig')

            # 统一为字符串比较，空值置空串
            df_curr_n = df_curr.astype(str).fillna('')
            df_base_n = df_base.astype(str).fillna('')

            # 检查结构变化
            added_cols = [c for c in df_curr_n.columns if c not in df_base_n.columns]
            removed_cols = [c for c in df_base_n.columns if c not in df_curr_n.columns]
            row_change = len(df_curr_n) - len(df_base_n)

            # 收集数据变更
            common_cols = [c for c in df_curr_n.columns if c in df_base_n.columns]
            min_rows = min(len(df_curr_n), len(df_base_n))
            changes = []

            # 比较公共行的数据变更
            for i in range(min_rows):
                for col in common_cols:
                    v_old = df_base_n.iat[i, df_base_n.columns.get_loc(col)]
                    v_new = df_curr_n.iat[i, df_curr_n.columns.get_loc(col)]
                    if v_old != v_new:
                        row_num = i + 2  # 行号+2(表头+索引)

                        # 使用统一的差异比较工具
                        diff_changes = self.compare_values_with_diff(v_old, v_new)
                        if diff_changes:
                            # 处理变更信息（统一6元素格式）
                            for change_type, arr_pos, old_item, new_item, _, arr_type in diff_changes:
                                if change_type == '删除':
                                    if arr_type == 'single':
                                        # 单个值变更
                                        changes.append((row_num, col, old_item, "删除"))
                                    else:
                                        # 数组项变更
                                        changes.append((row_num, f"{col}[{arr_pos}]", old_item, "删除"))
                                elif change_type == '新增':
                                    if arr_type == 'single':
                                        # 单个值变更
                                        changes.append((row_num, col, "新增", new_item))
                                    else:
                                        # 数组项变更
                                        changes.append((row_num, f"{col}[{arr_pos}]", "新增", new_item))
                                elif change_type == '替换':
                                    changes.append((row_num, col, old_item, new_item))
                        else:
                            # 如果没有检测到变更，但值确实不同，显示整体变更
                            changes.append((row_num, col, v_old, v_new))

            # 处理新增行
            if len(df_curr_n) > len(df_base_n):
                for i in range(len(df_base_n), len(df_curr_n)):
                    row_num = i + 2  # 行号+2(表头+索引)
                    for col in common_cols:
                        v_new = df_curr_n.iat[i, df_curr_n.columns.get_loc(col)]
                        # 跳过空值
                        if pd.notna(v_new) and str(v_new).strip() and str(v_new) != 'nan':
                            # 使用统一的数组解析
                            array_items, arr_type = self.parse_array_value(v_new)
                            if arr_type == 'single':
                                # 单个值
                                changes.append((row_num, col, "新增", v_new))
                            else:
                                # 数组值，拆包为单个项目
                                for idx, item in enumerate(array_items):
                                    if item.strip():  # 跳过空项
                                        changes.append((row_num, f"{col}[{idx}]", "新增", item))

            # 处理删除行
            if len(df_base_n) > len(df_curr_n):
                for i in range(len(df_curr_n), len(df_base_n)):
                    row_num = i + 2  # 行号+2(表头+索引)
                    for col in common_cols:
                        v_old = df_base_n.iat[i, df_base_n.columns.get_loc(col)]
                        # 跳过空值
                        if pd.notna(v_old) and str(v_old).strip() and str(v_old) != 'nan':
                            # 使用统一的数组解析
                            array_items, arr_type = self.parse_array_value(v_old)
                            if arr_type == 'single':
                                # 单个值
                                changes.append((row_num, col, v_old, "删除"))
                            else:
                                # 数组值，拆包为单个项目
                                for idx, item in enumerate(array_items):
                                    if item.strip():  # 跳过空项
                                        changes.append((row_num, f"{col}[{idx}]", item, "删除"))

            # 打印摘要
            if added_cols or removed_cols or row_change != 0:
                print(f"  结构变更:")
                if added_cols:
                    print(f"    新增列: {', '.join(added_cols)}")
                if removed_cols:
                    print(f"    删除列: {', '.join(removed_cols)}")
                if row_change != 0:
                    print(f"    行数变化: {'+' if row_change > 0 else ''}{row_change}")

            if changes:
                print(f"  数据变更 ({len(changes)} 处):")

                # 对变更进行排序：先按行号，再按列名（去掉数组索引部分）
                def sort_key(change):
                    row, col, _, _ = change
                    # 提取基础列名（去掉[索引]部分）
                    base_col = col.split('[')[0] if '[' in col else col
                    return (row, base_col, col)  # 行号、基础列名、完整列名

                sorted_changes = sorted(changes, key=sort_key)

                # 显示所有变更，使用对齐格式
                display_changes = sorted_changes

                # 计算各列的最大宽度
                max_row_col_width = 0
                max_old_width = 0
                formatted_changes = []

                for row, col, old_val, new_val in display_changes:
                    row_col = f"行{row} [{col}]"
                    old_display = self.truncate_text(old_val, 35)
                    new_display = self.truncate_text(new_val, 35)

                    max_row_col_width = max(max_row_col_width, self.get_display_width(row_col))
                    max_old_width = max(max_old_width, self.get_display_width(old_display))

                    formatted_changes.append((row_col, old_display, new_display))

                # 输出对齐的表格
                for row_col, old_display, new_display in formatted_changes:
                    # 计算需要的空格数来对齐
                    row_col_padding = max_row_col_width - self.get_display_width(row_col)
                    old_padding = max_old_width - self.get_display_width(old_display)

                    print(f"    {row_col}{' ' * row_col_padding}  {old_display}{' ' * old_padding}  →  {new_display}")
            else:
                if not (added_cols or removed_cols or row_change != 0):
                    print(f"  无变更")

        except Exception as e:
            print(f"  比对失败: {e}")

    def convert(self, command):
        """执行转换操作"""
        try:
            # 解析命令
            filename, sheet_name = self.parse_command(command)
            print(f"解析命令: 文件名='{filename}', 工作表='{sheet_name}'")

            # 查找Excel文件
            file_path = self.find_excel_file(filename)
            if not file_path:
                raise FileNotFoundError(f"在目录 '{self.target_folder}' 中未找到文件 '{filename}'")

            print(f"找到文件: {file_path}")

            # 读取指定工作表
            print(f"正在读取工作表 '{sheet_name}'...")
            df = self.read_excel_sheet(file_path, sheet_name)

            print(f"成功读取数据: {len(df)} 行, {len(df.columns)} 列")

            # 构建当前表和工作表的标识
            base_filename = Path(filename).stem  # 去掉扩展名
            current_table_sheet = f"{base_filename}[{sheet_name}]"

            # 预预处理DataFrame，处理自定义字段关联
            df_pre_processed = self.pre_preprocess_dataframe(df, current_table_sheet)

            # 预处理DataFrame，查找并替换t_*字符串
            df_processed = self.preprocess_dataframe(df_pre_processed)

            # 生成输出文件名（去掉.xls扩展名）
            base_filename = Path(filename).stem  # 去掉扩展名
            output_filename = f"{base_filename}[{sheet_name}].csv"

            # 保存为CSV
            print(f"正在保存为CSV文件: {output_filename}")
            output_path = self.save_to_csv(df_processed, output_filename)

            # 首次导出时保存基线备份（如已存在则不覆盖）
            base_output_path = self.base_folder / output_filename
            try:
                if not base_output_path.exists():
                    shutil.copyfile(output_path, base_output_path)
                    print(f"保存基线备份: {base_output_path.name}")
                else:
                    print(f"基线备份已存在: {base_output_path.name}")
            except Exception as be:
                print(f"保存基线备份失败: {be}")

            print(f"输出: {output_path.name} ({len(df_processed)}行 x {len(df_processed.columns)}列)")

        except Exception as e:
            print(f"❌ 转换失败: {str(e)}")
            return False

        return True

    def process_csv_content(self, csv_content):
        """处理CSV内容，将各种{中文}格式还原为原始格式"""
        import re

        # 正则表达式匹配 t_*{中文} 格式
        t_pattern = r't_([a-zA-Z0-9_]+)\{[^}]*\}'
        # 替换为 t_* 格式
        processed_content = re.sub(t_pattern, r't_\1', csv_content)

        # 正则表达式匹配 数字{中文} 格式（用于ID关联）
        id_pattern = r'(\d+)\{[^}]*\}'
        # 替换为纯数字格式
        processed_content = re.sub(id_pattern, r'\1', processed_content)

        return processed_content

    def extract_and_update_language_texts(self, new_item):
        """从变更项中提取t_*{中文}格式并更新语言表

        Args:
            new_item: 变更的新值，可能包含t_*{中文}格式

        Returns:
            bool: 是否有语言文本需要更新
        """
        if not new_item or not isinstance(new_item, str):
            return False

        import re

        # 正则表达式匹配 t_*{中文} 格式
        pattern = r't_([a-zA-Z0-9_]+)\{([^}]+)\}'
        matches = re.findall(pattern, new_item)

        if not matches:
            return False

        print(f"发现 {len(matches)} 个语言文本更新需求")

        # 导入go.py模块
        try:
            import sys
            sys.path.append(str(Path.cwd()))
            from go import ExcelTextReplacer
        except Exception as e:
            print(f"导入go.py模块失败: {str(e)}")
            return False

        # 创建文本替换器实例
        replacer = ExcelTextReplacer({})

        success_count = 0
        for t_id_part, chinese_text in matches:
            t_full_id = f"t_{t_id_part}"
            print(f"处理语言文本: {t_full_id} -> {chinese_text}")

            # 直接调用go.py的语言文本更新方法
            # go.py会自动处理查找、更新或新增逻辑
            if replacer.update_language_text_by_id(t_full_id, chinese_text):
                success_count += 1
            else:
                print(f"  更新失败: {t_full_id}")

        print(f"语言文本更新完成: {success_count}/{len(matches)}")
        return success_count > 0



    def apply_language_text_changes(self, changes):
        """应用语言文本变更到原文件

        Args:
            changes: 变更记录列表，格式为 (row_num, col, old_item, new_item, arr_pos, arr_type)
        """
        print("开始应用语言文本变更...")

        processed_count = 0
        for change in changes:
            if len(change) >= 4:  # 确保有足够的参数
                _, _, _, new_item = change[:4]  # 只使用new_item

                # 只处理新增和修改的项目
                if new_item and new_item != "删除":
                    if self.extract_and_update_language_texts(new_item):
                        processed_count += 1

        print(f"语言文本变更应用完成，处理了 {processed_count} 个变更项")

    def sync_changes_to_original_files(self, csv_file_path):
        """将CSV变更同步到原始文件

        这是你提出的完整方案的实现：
        1. 获取变更记录（包含数组位置和格式信息）
        2. 先处理语言文本关联（t_*{中文}格式）
        3. 再处理数据同步到原文件

        Args:
            csv_file_path: CSV文件路径
        """
        try:
            print("开始同步变更到原始文件...")

            # 1. 获取增强的变更记录
            enhanced_changes = self.get_enhanced_changes_with_baseline(csv_file_path)

            if not enhanced_changes:
                print("没有检测到变更，跳过同步")
                return True

            print(f"检测到 {len(enhanced_changes)} 个变更项")

            # 2. 先处理语言文本关联
            print("\n步骤1: 处理语言文本关联...")
            self.apply_language_text_changes(enhanced_changes)

            # 3. 再处理数据同步到原文件
            print("\n步骤2: 处理数据同步...")
            self.apply_data_changes_to_original_files(enhanced_changes, csv_file_path)

            # 4. 重新生成CSV和更新基线
            print("\n步骤3: 更新CSV和基线...")
            self.refresh_csv_and_baseline_after_sync(csv_file_path)

            print("变更同步完成！")
            return True

        except Exception as e:
            print(f"同步变更时出错: {str(e)}")
            return False

    def refresh_csv_and_baseline_after_sync(self, csv_file_path):
        """同步完成后重新生成CSV和更新基线

        这样可以看到改后的真实样子，也为下次比对做好准备

        Args:
            csv_file_path: 当前CSV文件路径
        """
        try:
            # 从CSV文件名解析出Excel文件信息
            csv_file = Path(csv_file_path)
            csv_name = csv_file.stem  # 例如: "hero[hero]"

            # 解析文件名和工作表名
            if '[' in csv_name and ']' in csv_name:
                excel_name = csv_name.split('[')[0]  # "hero"
                sheet_name = csv_name.split('[')[1].rstrip(']')  # "hero"
            else:
                print(f"⚠️ 无法解析CSV文件名格式: {csv_name}")
                return False

            # 构建Excel文件路径
            excel_file_path = Path(self.target_folder) / f"{excel_name}.xls"
            if not excel_file_path.exists():
                excel_file_path = Path(self.target_folder) / f"{excel_name}.xlsx"
                if not excel_file_path.exists():
                    print(f"⚠️ 找不到对应的Excel文件: {excel_name}.xls 或 {excel_name}.xlsx")
                    return False

            print(f"  重新生成CSV: {excel_file_path} → {csv_file_path}")

            # 重新调用完整的转换流程（包含三步处理）
            # 这样生成的CSV会包含注释，Excel是正式配置不带注释
            command = f"{excel_name}[{sheet_name}]"
            success = self.convert(command)
            if not success:
                print(f"  ❌ 重新生成CSV失败")
                return False

            # 更新基线文件
            base_csv_path = self.base_folder / csv_file.name
            base_csv_path.parent.mkdir(parents=True, exist_ok=True)

            print(f"  更新基线: {csv_file_path} → {base_csv_path}")

            # 复制当前CSV到基线文件夹
            import shutil
            shutil.copy2(csv_file_path, base_csv_path)

            print(f"  ✅ CSV和基线更新完成")
            return True

        except Exception as e:
            print(f"  ❌ 更新CSV和基线时出错: {str(e)}")
            return False

    def get_enhanced_changes_with_baseline(self, csv_file_path):
        """获取增强的变更记录（包含数组位置和格式信息）

        Returns:
            list: 增强的变更记录列表，格式为 (row_num, col, old_item, new_item, arr_pos, arr_type)
        """
        try:
            csv_path = Path(csv_file_path)
            base_path = self.base_folder / csv_path.name

            if not base_path.exists():
                print("无基线备份，无法获取变更记录")
                return []

            # 读取当前与基线
            df_curr = pd.read_csv(csv_path, encoding='utf-8-sig')
            df_base = pd.read_csv(base_path, encoding='utf-8-sig')

            # 统一为字符串比较，空值置空串
            df_curr_n = df_curr.astype(str).fillna('')
            df_base_n = df_base.astype(str).fillna('')

            # 收集增强的变更记录
            enhanced_changes = []
            common_cols = [c for c in df_curr_n.columns if c in df_base_n.columns]
            min_rows = min(len(df_curr_n), len(df_base_n))

            # 比较公共行的数据变更
            for i in range(min_rows):
                for col in common_cols:
                    v_old = df_base_n.iat[i, df_base_n.columns.get_loc(col)]
                    v_new = df_curr_n.iat[i, df_curr_n.columns.get_loc(col)]
                    if v_old != v_new:
                        row_num = i + 2  # 行号+2(表头+索引)

                        # 使用统一的差异比较工具
                        diff_changes = self.compare_values_with_diff(v_old, v_new)
                        if diff_changes:
                            enhanced_changes.extend([
                                (row_num, col, old_item, new_item, arr_pos, arr_type)
                                for _, arr_pos, old_item, new_item, _, arr_type in diff_changes
                            ])

            return enhanced_changes

        except Exception as e:
            print(f"获取变更记录时出错: {str(e)}")
            return []

    def apply_data_changes_to_original_files(self, enhanced_changes, csv_file_path):
        """将数据变更精确应用到原始Excel文件

        Args:
            enhanced_changes: 增强的变更记录列表
            csv_file_path: CSV文件路径，用于确定目标Excel文件
        """
        if not enhanced_changes:
            print("没有数据变更需要应用")
            return True

        # 解析CSV文件名获取目标Excel信息
        csv_path = Path(csv_file_path)
        csv_filename = csv_path.stem

        if '[' not in csv_filename or ']' not in csv_filename:
            print(f"CSV文件名格式错误: {csv_filename}")
            return False

        file_part, sheet_part = csv_filename.split('[', 1)
        sheet_name = sheet_part.rstrip(']')
        excel_filename = f"{file_part.strip()}.xls"

        # 查找目标Excel文件
        excel_file_path = self.find_excel_file(excel_filename)
        if not excel_file_path:
            print(f"未找到目标Excel文件: {excel_filename}")
            return False

        print(f"应用变更到: {excel_file_path.name}[{sheet_name}]")

        # 按行列分组处理变更（同一个单元格的所有变更一起处理）
        changes_by_cell = self._group_changes_by_cell(enhanced_changes)

        success_count = 0
        total_count = len(changes_by_cell)

        # 处理每个单元格的变更
        for (row_num, col), cell_changes in changes_by_cell.items():
            print(f"  处理单元格 行{row_num} 列{col}: {len(cell_changes)} 个变更")

            if self._apply_cell_changes_to_excel(
                excel_file_path, sheet_name, row_num, col, cell_changes
            ):
                success_count += 1

        print(f"数据同步完成: {success_count}/{total_count} 个单元格已更新")
        return success_count == total_count

    def _group_changes_by_cell(self, changes):
        """按单元格分组变更"""
        cell_groups = {}

        for change in changes:
            if len(change) >= 6:
                row_num, col, old_item, new_item, arr_pos, arr_type = change
                cell_key = (row_num, col)

                if cell_key not in cell_groups:
                    cell_groups[cell_key] = []

                cell_groups[cell_key].append({
                    'old_item': old_item,
                    'new_item': new_item,
                    'arr_pos': arr_pos,
                    'arr_type': arr_type
                })

        return cell_groups

    def _apply_cell_changes_to_excel(self, excel_file_path, sheet_name, row_num, col, cell_changes):
        """将单元格的所有变更一次性应用到Excel文件"""
        try:
            # 导入go.py模块
            import sys
            sys.path.append(str(Path.cwd()))
            from go import ExcelTextReplacer

            # 创建替换器实例
            replacer = ExcelTextReplacer({})

            # 调用go.py的方法进行单元格级别的更新
            return replacer.update_cell_with_multiple_changes(
                str(excel_file_path), sheet_name, row_num, col, cell_changes
            )

        except Exception as e:
            print(f"应用单元格变更时出错: {str(e)}")
            return False

    def _group_changes_by_type(self, changes):
        """按变更类型分组"""
        groups = {'新增': [], '删除': [], '替换': []}

        for change in changes:
            # 变更记录格式: (row_num, col, old_item, new_item, arr_pos, arr_type)
            # 需要从变更记录中推断变更类型
            if len(change) >= 4:
                row_num, col, old_item, new_item, arr_pos, arr_type = change

                if old_item and new_item:
                    change_type = '替换'
                elif new_item and not old_item:
                    change_type = '新增'
                elif old_item and not new_item:
                    change_type = '删除'
                else:
                    continue  # 跳过无效变更

                # 重新构造变更记录，包含推断的类型
                enhanced_change = (row_num, col, old_item, new_item, arr_pos, arr_type)
                groups[change_type].append(enhanced_change)

        return {k: v for k, v in groups.items() if v}  # 只返回非空组

    def _apply_single_change_to_excel(self, excel_file_path, sheet_name, row_num, col, new_value, arr_pos, arr_type, change_type):
        """将单个变更应用到Excel文件

        Args:
            excel_file_path: Excel文件路径
            sheet_name: 工作表名
            row_num: 行号（1基索引，包含表头）
            col: 列名
            new_value: 新值
            arr_pos: 数组位置
            arr_type: 数组类型
            change_type: 变更类型

        Returns:
            bool: 是否成功
        """
        try:
            # 导入go.py模块
            import sys
            sys.path.append(str(Path.cwd()))
            from go import ExcelTextReplacer

            # 创建替换器实例
            replacer = ExcelTextReplacer({})

            # 调用go.py的方法进行精确更新
            return replacer.update_cell_value_precisely(
                str(excel_file_path), sheet_name, row_num, col,
                new_value, arr_pos, arr_type, change_type
            )

        except Exception as e:
            print(f"应用变更时出错: {str(e)}")
            return False

    def write_csv_to_excel(self, csv_file_path, excel_file_path, sheet_name):
        """将CSV文件内容写回到Excel文件的指定工作表"""
        try:
            # 读取CSV文件
            df = pd.read_csv(csv_file_path, encoding='utf-8-sig')

            # 处理CSV内容，将t_*{中文}格式还原为t_*格式
            for col in df.columns:
                df[col] = df[col].astype(str).apply(lambda x: self.process_csv_content(x) if pd.notna(x) and x != 'nan' else x)

            # 将'nan'字符串转换回NaN
            df = df.replace('nan', pd.NA)

            # 检查Excel文件类型并写入
            file_extension = Path(excel_file_path).suffix.lower()

            if file_extension == '.xlsx':
                self.write_to_xlsx(df, excel_file_path, sheet_name)
            elif file_extension == '.xls':
                self.write_to_xls(df, excel_file_path, sheet_name)
            else:
                raise ValueError(f"不支持的Excel文件格式: {file_extension}")

        except Exception as e:
            raise Exception(f"写入Excel文件失败: {str(e)}")

    def write_to_xlsx(self, df, excel_file_path, sheet_name):
        """将DataFrame写入.xlsx文件的指定工作表"""
        try:
            # 检查文件是否存在
            if Path(excel_file_path).exists():
                # 文件存在，读取现有工作簿
                workbook = openpyxl.load_workbook(excel_file_path)

                # 记录原始工作表顺序
                original_sheet_names = workbook.sheetnames.copy()
                target_sheet_index = -1

                # 如果工作表存在，记录其位置并删除
                if sheet_name in workbook.sheetnames:
                    target_sheet_index = original_sheet_names.index(sheet_name)
                    del workbook[sheet_name]

                # 创建新的工作表
                worksheet = workbook.create_sheet(sheet_name)

                # 如果找到了原始位置，将工作表移动到正确位置
                if target_sheet_index != -1:
                    # 将新工作表移动到原始位置
                    workbook.move_sheet(worksheet, target_sheet_index)

            else:
                # 文件不存在，创建新工作簿
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = sheet_name

            # 写入列标题
            for col_idx, column_name in enumerate(df.columns, 1):
                worksheet.cell(row=1, column=col_idx, value=column_name)

            # 写入数据
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                for col_idx, value in enumerate(row, 1):
                    # 处理NaN值
                    if pd.isna(value):
                        cell_value = None
                    else:
                        cell_value = value
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

            # 保存文件
            workbook.save(excel_file_path)
            workbook.close()

        except Exception as e:
            raise Exception(f"写入.xlsx文件失败: {str(e)}")

    def write_to_xls(self, df, excel_file_path, sheet_name):
        """将DataFrame写入.xls文件的指定工作表"""
        try:
            # 对于.xls文件，我们需要重新创建整个文件
            # 因为xlwt不支持修改现有文件

            # 如果原文件存在，先读取所有工作表（按原始顺序）
            existing_sheets = []  # 使用列表保持顺序
            target_sheet_index = -1  # 目标工作表的原始位置

            if Path(excel_file_path).exists():
                try:
                    old_workbook = xlrd.open_workbook(excel_file_path)
                    for sheet_idx in range(old_workbook.nsheets):
                        old_sheet = old_workbook.sheet_by_index(sheet_idx)
                        old_sheet_name = old_sheet.name

                        if old_sheet_name == sheet_name:
                            # 记录目标工作表的位置
                            target_sheet_index = sheet_idx
                            # 为目标工作表预留位置
                            existing_sheets.append((old_sheet_name, None))
                        else:
                            # 保存其他工作表的数据
                            sheet_data = []
                            for row_idx in range(old_sheet.nrows):
                                row_data = []
                                for col_idx in range(old_sheet.ncols):
                                    cell_value = old_sheet.cell_value(row_idx, col_idx)
                                    row_data.append(cell_value)
                                sheet_data.append(row_data)
                            existing_sheets.append((old_sheet_name, sheet_data))
                except Exception as e:
                    print(f"警告: 读取原有.xls文件时出错: {str(e)}")

            # 如果没有找到目标工作表，添加到末尾
            if target_sheet_index == -1:
                existing_sheets.append((sheet_name, None))

            # 创建新的工作簿
            new_workbook = xlwt.Workbook()

            # 按原始顺序添加工作表
            for sheet_name_in_order, sheet_data in existing_sheets:
                if sheet_data is None:
                    # 这是目标工作表，写入新数据
                    worksheet = new_workbook.add_sheet(sheet_name_in_order)

                    # 写入列标题
                    for col_idx, column_name in enumerate(df.columns):
                        worksheet.write(0, col_idx, column_name)

                    # 写入数据
                    for row_idx, (_, row) in enumerate(df.iterrows(), 1):
                        for col_idx, value in enumerate(row):
                            # 处理NaN值
                            if pd.isna(value):
                                cell_value = ""
                            else:
                                cell_value = value
                            worksheet.write(row_idx, col_idx, cell_value)
                else:
                    # 这是现有工作表，复制原数据
                    worksheet = new_workbook.add_sheet(sheet_name_in_order)
                    for row_idx, row_data in enumerate(sheet_data):
                        for col_idx, cell_value in enumerate(row_data):
                            worksheet.write(row_idx, col_idx, cell_value)

            # 保存文件
            new_workbook.save(excel_file_path)

        except Exception as e:
            raise Exception(f"写入.xls文件失败: {str(e)}")

    def get_sheet_names(self, excel_file_path):
        """获取Excel文件的工作表名称列表"""
        try:
            file_extension = Path(excel_file_path).suffix.lower()

            if file_extension == '.xlsx':
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_names = workbook.sheetnames.copy()
                workbook.close()
                return sheet_names
            elif file_extension == '.xls':
                workbook = xlrd.open_workbook(excel_file_path)
                sheet_names = workbook.sheet_names()
                return sheet_names
            else:
                return []
        except Exception as e:
            print(f"警告: 读取工作表名称时出错: {str(e)}")
            return []

    def update_excel_from_csv(self):
        """遍历xls文件夹中的CSV文件，将其内容写回到对应的Excel文件"""
        try:


            # 查找所有CSV文件
            csv_files = list(self.output_folder.glob("*.csv"))

            if not csv_files:
                print(f"在文件夹 '{self.output_folder}' 中未找到CSV文件")
                return False

            if len(csv_files) == 1:
                print(f"找到 1 个CSV文件: {csv_files[0].name}")
            else:
                print(f"找到 {len(csv_files)} 个CSV文件:")
                for csv_file in csv_files:
                    print(f"  {csv_file.name}")
            print()

            success_count = 0

            # 处理每个CSV文件
            for csv_file in csv_files:
                try:
                    # 解析CSV文件名
                    csv_filename = csv_file.stem
                    if '[' not in csv_filename or ']' not in csv_filename:
                        print(f"跳过 {csv_file.name} (格式错误)")
                        continue

                    file_part, sheet_part = csv_filename.split('[', 1)
                    sheet_name = sheet_part.rstrip(']')
                    excel_filename = f"{file_part.strip()}.xls"

                    print(f"处理 {csv_file.name} → {excel_filename}[{sheet_name}]")

                    # 查找Excel文件
                    excel_file_path = self.find_excel_file(excel_filename)
                    if not excel_file_path:
                        print(f"  错误: 未找到 {excel_filename}")
                        continue

                    # 第1步：找到差异
                    print(f"  第1步: 分析变更差异...")
                    self.show_diff_with_baseline(csv_file)

                    # 第2步：智能同步变更到原文件
                    print(f"  第2步: 智能同步变更...")
                    sync_success = self.sync_changes_to_original_files(csv_file)

                    if sync_success:
                        print(f"  ✅ 智能同步完成")
                    else:
                        print(f"  ⚠️ 智能同步失败，使用传统方法...")
                        # 如果智能同步失败，回退到传统方法
                        self.write_csv_to_excel(csv_file, excel_file_path, sheet_name)
                        print(f"  完成传统写入")

                    success_count += 1

                except Exception as e:
                    print(f"  错误: {str(e)}")

                print()

            # 结果摘要
            if success_count == len(csv_files):
                print(f"✅ 全部完成 ({success_count}/{len(csv_files)})")
            else:
                print(f"⚠️ 部分完成 ({success_count}/{len(csv_files)})")
            return success_count > 0

        except Exception as e:
            print(f"❌ 更新失败: {str(e)}")
            return False

def main():
    """主函数"""
    # 创建转换器实例
    converter = ExcelToCSVConverter(TARGET_FOLDER, OUTPUT_FOLDER)

    # 检查命令行参数
    if len(sys.argv) == 1:
        # 没有参数，执行CSV到Excel的更新操作
        print("Excel更新工具")
        print(f"目标文件夹: {TARGET_FOLDER}")
        print(f"CSV文件夹: {converter.output_folder}")
        print(f"基线文件夹: {converter.base_folder}")
        print()

        success = converter.update_excel_from_csv()

        # success 结果已在 update_excel_from_csv 中显示

    elif len(sys.argv) == 2:
        # 有一个参数，执行Excel到CSV的导出操作
        command = sys.argv[1]

        print("Excel导出工具")
        print(f"目标文件夹: {TARGET_FOLDER}")
        print(f"输出文件夹: {Path.cwd()}/{OUTPUT_FOLDER}")
        print(f"执行: {command}")
        print()

        success = converter.convert(command)

        if success:
            print("✅ 导出完成")
        else:
            print("❌ 导出失败")
    else:
        # 参数错误
        print("Excel配置工具")
        print("="*50)
        print("使用方法:")
        print("1. 导出Excel工作表为CSV:")
        print("   py config.py filename[sheetname]")
        print("   示例: py config.py hero[hero]")
        print()
        print("2. 将CSV文件写回Excel:")
        print("   py config.py")
        print("   (无参数，自动处理xls文件夹中的所有CSV文件)")
        print("="*50)

if __name__ == "__main__":
    main()
